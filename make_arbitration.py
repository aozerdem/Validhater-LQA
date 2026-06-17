"""
make_arbitration.py
Generate a focused arbitration workbook for the 5 calibration disagreements
(goldset rows 2, 5, 13, 39, 40) so the Language Lead can give a final verdict.

These are the FALSE_PASS rows from calibration_report_remap_20260612_130451:
the AI judged the MT publishable but the goldset LL had marked it "No".
Row 13 is included for confirmation (already relabelled No->Yes pending sign-off).
"""

from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment

# Each row: goldset data + AI evaluation + the specific question + Claude's tentative read
ROWS = [
    {
        "row": 2,
        "source": "Xunulyn Classic Unisex Baseball Cap Adjustable Cute Puppy Flower Cartoon Hand Drawn can be Used Print Kids wear Fashion Design Baby sho",
        "mt": "Xunulyn Classic Unisex baseballcap Justerbar søt valpeblomst Tegneserie Håndtegnet kan brukes Trykk Barneklær Motedesign Baby sho",
        "ll_verdict": "No",
        "ll_reason": "\"Cute Puppy Flower\" should not be translated as \"søt valpeblomst\". It should be seen as two segments: \"Cute Puppy\" = \"Søt valp\", and \"Flower\" = \"Blomst\" as its own segment.",
        "ai_sev": "WARN", "ai_score": 95, "ai_cat": "accuracy:mistranslation",
        "ai_reason": "Source uses 'Baseball Cap' which per LL guidance should be 'baseballcaps', not 'baseballcap'. Otherwise the segment mirrors the (nonsensical) source structure appropriately.",
        "question": "The source is an unpunctuated keyword-stuffed title. Is rendering 'Cute Puppy Flower' as the compound 'søt valpeblomst' a PUBLISH-BLOCKING error, or acceptable given the garbled source? (AI treated it as advisory WARN, i.e. publishable.)",
        "claude_read": "Gray zone. On unpunctuated garbage, the compound reading is defensible — I lean publishable/advisory, but not confident. Your call decides it.",
    },
    {
        "row": 5,
        "source": "Quality is great and the material feels very sturdy.",
        "mt": "Kvaliteten er flott og materialet føles veldig solid.",
        "ll_verdict": "No",
        "ll_reason": "Based on context it is better to use \"god\" as a translation for \"great\", as this is more idiomatically correct.",
        "ai_sev": "OK", "ai_score": 98, "ai_cat": "no-error",
        "ai_reason": "(scored clean — no error flagged)",
        "question": "Is 'Kvaliteten er flott' genuinely NOT publishable, or is 'god' a preference? Does Amazon house style / prior escalation mandate 'god' over 'flott' here?",
        "claude_read": "I lean AI: 'Kvaliteten er flott' reads as natural, idiomatic marketing Norwegian. But you may have client-style context I lack — this is exactly where I could be wrong.",
    },
    {
        "row": 13,
        "source": "11.5in X 7.5in X 5.7in",
        "mt": "11,5 tommer X 7,5 tommer X 5,7 tommer",
        "ll_verdict": "No  (already relabelled -> Yes, pending your confirmation)",
        "ll_reason": "Measurements should be localized depending on context. (As inch is used in some cases in Norwegian, I have not edited the post-edited translation)",
        "ai_sev": "OK", "ai_score": 98, "ai_cat": "no-error",
        "ai_reason": "(scored clean — no error flagged)",
        "question": "CONFIRMATION: the recorded LL note says inch is acceptable and the MT was NOT edited (pe field empty, no error spans). The 'No' label appears to contradict the note. Should this be 'Yes' (publishable)?",
        "claude_read": "Confident this is a label artifact: every field except the verdict says clean. Relabelled No->Yes in the goldset with audit note — please confirm or revert.",
    },
    {
        "row": 39,
        "source": "Gold,Red,Blue,Titanium,Black?blue1 Fitment: for Ducati 1098 S Tricolor 2007 2008 ( Please Ensure This Part Fits For Your Motorcycle Before order ) package:",
        "mt": "Gull, rød, blå, titan, svart? blue1 Montering: for Ducati 1098 S Tricolor 2007 2008 (Forsikre deg om at denne delen passer til motorsykkelen din før bestilling) pakke:",
        "ll_verdict": "No",
        "ll_reason": "Term has not been translated",
        "ai_sev": "OK", "ai_score": 98, "ai_cat": "no-error",
        "ai_reason": "(scored clean — treated 'blue1' as a broken/corrupted source token to mirror verbatim)",
        "question": "The source 'Black?blue1' is corrupted. Should 'blue1' be (a) mirrored verbatim as a broken token, or (b) translated to 'blå1'? The AI mirrored it; the LL flagged it as untranslated.",
        "claude_read": "Gray zone. 'blue1' comes from corrupted 'Black?blue1', so mirroring is defensible per our broken-token rule — but a WARN flag on 'blue1' would also be reasonable. Your call.",
    },
    {
        "row": 40,
        "source": "1 brake and clutch lever",
        "mt": "1 brems- og koblingsspak",
        "ll_verdict": "No",
        "ll_reason": "the correct terminology has not been used. Segment should therefore be edited",
        "ai_sev": "OK", "ai_score": 99, "ai_cat": "no-error",
        "ai_reason": "(scored clean — no error flagged)",
        "question": "Is 'brems- og koblingsspak' wrong terminology? Specifically: should the truncated compound be 'bremse- og koblingsspak' (full stem)? And is 'koblingsspak' the right term for 'clutch lever', or should it be 'clutchspak'?",
        "claude_read": "I believe the LL is RIGHT and the AI MISSED this: 'brems-' should be 'bremse-' (stem of 'bremsespak'). Flagged here so the goldset 'No' is NOT changed — this looks like a genuine tool miss, not a label error.",
    },
]


def main():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Arbitration_5rows"

    title_font = Font(bold=True, size=12)
    hdr_fill = PatternFill("solid", fgColor="2E4057")
    hdr_font = Font(bold=True, color="FFFFFF")
    ask_fill = PatternFill("solid", fgColor="FFF2CC")   # pale yellow = please fill in
    read_fill = PatternFill("solid", fgColor="EAEAEA")  # grey = FYI only
    wrap = Alignment(wrap_text=True, vertical="top")

    # Intro / instructions
    ws.merge_cells("A1:K1")
    ws["A1"] = ("Arbitration — 5 calibration disagreements (Amazon TDT EN-GB -> NB-NO). "
                "The automated checker judged each MT publishable; the goldset LL had marked it 'No'. "
                "Please give your final verdict in the two yellow columns. The grey 'Claude's tentative read' "
                "column is FYI only — please judge independently.")
    ws["A1"].alignment = Alignment(wrap_text=True, vertical="top")
    ws["A1"].font = title_font
    ws.row_dimensions[1].height = 60

    headers = [
        "Goldset row", "Source (EN-GB)", "MT target (NB-NO)",
        "LL verdict (goldset)", "LL reason (goldset)",
        "AI severity", "AI score", "AI category",
        "Question for you",
        ">> YOUR VERDICT (AI right / LL right / Both partly)",
        ">> YOUR notes / correct rendering",
    ]
    ws.append(headers)
    for cell in ws[2]:
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 44

    for r in ROWS:
        ws.append([
            r["row"], r["source"], r["mt"],
            r["ll_verdict"], r["ll_reason"],
            r["ai_sev"], r["ai_score"], r["ai_cat"],
            r["question"], "", "",
        ])
        rownum = ws.max_row
        for cell in ws[rownum]:
            cell.alignment = wrap
        ws.cell(rownum, 10).fill = ask_fill
        ws.cell(rownum, 11).fill = ask_fill
        ws.row_dimensions[rownum].height = 110

    # A separate, clearly-secondary FYI block below the table
    ws.append([])
    ws.append(["Claude's tentative read (FYI ONLY — please do not be biased; judge independently):"])
    ws.cell(ws.max_row, 1).font = Font(bold=True, italic=True)
    for r in ROWS:
        ws.append([f"Row {r['row']}", r["claude_read"]])
        ws.cell(ws.max_row, 1).font = Font(bold=True)
        ws.cell(ws.max_row, 2).fill = read_fill
        ws.cell(ws.max_row, 2).alignment = wrap
        ws.merge_cells(start_row=ws.max_row, start_column=2, end_row=ws.max_row, end_column=9)
        ws.row_dimensions[ws.max_row].height = 44

    widths = [12, 50, 50, 22, 45, 10, 8, 22, 55, 28, 40]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    ws.freeze_panes = "A3"

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    out = Path(__file__).parent / f"arbitration_5rows_{timestamp}.xlsx"
    wb.save(out)
    print(f"Saved: {out}")


if __name__ == "__main__":
    main()
