"""
va_evaluator.py
Amazon TDT EN-GB > NB-NO — Validation QA Evaluator
------------------------------------------------------
Checkpoint 1: Excel reader + stub output
Checkpoint 2: AI evaluator (single-segment, calibration)
Checkpoint 3: Batch + aggregation
Checkpoint 4: Excel output

Usage:
    python va_evaluator.py file1.xlsx [file2.xlsx ...]

Output:
    va_report_<timestamp>.xlsx   (separate file, source exports untouched)
"""

import sys
import json
import re
import time
import random
import threading
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime
from pathlib import Path
import os
import boto3
from botocore.exceptions import ClientError, ReadTimeoutError, ConnectTimeoutError, EndpointConnectionError
import openpyxl

# ─────────────────────────────────────────────
# COLUMN INDICES (0-based) in Segments sheet
# ─────────────────────────────────────────────
COL_SOURCE          = 4   # E  SourceSegment
COL_MT_TARGET       = 5   # F  OriginalTargetSegment (MT output validator signed off on)
COL_ORIGIN_TYPE     = 10  # K  SegmentOriginType
COL_VALIDATOR_NAME  = 15  # P  ValidationResource
COL_VALIDATOR_ID    = 16  # Q  ValidatorGalileoID
COL_PUBLISH         = 20  # U  Publish  ("Yes" / "No")
COL_SEGMENT_ID      = 42  # AS SegmentID (source filename)
COL_FILE_NAME       = 41  # AR FileName

# ── PEQA: post-editing handoff export (different sheet layout) ──
COL_PE_SOURCE       = 4   # E  SourceSegment
COL_PE_MT           = 5   # F  OriginalTargetSegment (raw MT, context only)
COL_PE_RESOURCE     = 14  # O  Post-EditingResource (the linguist who post-edited)
COL_PE_GALILEO_ID   = 15  # P  GalileoID
COL_PE_TARGET       = 16  # Q  Post-EditingTargetSegment (final output — THIS is scored)
COL_PE_TER          = 18  # S  Post-EditingTER (edit distance from MT)
COL_PE_FILE_NAME    = 33  # AH FileName
COL_PE_SEGMENT_ID   = 34  # AI SegmentID

# ─────────────────────────────────────────────
# SCORING THRESHOLDS (mirrors Amazon grading)
# ─────────────────────────────────────────────
SCORE_OK   = 98   # >= 98  → OK   (Amazon routinely scores 97-99 on passing files)
SCORE_WARN = 95   # >= 95  → WARN (below this = Fail in Amazon's rubric)
              #  < 95  → FAIL

# ─────────────────────────────────────────────
# ERROR CATEGORIES (Amazon scorecard taxonomy)
# ─────────────────────────────────────────────
ERROR_CATEGORIES = [
    "accuracy:mistranslation",
    "accuracy:omission",
    "accuracy:addition",
    "accuracy:untranslated",
    "fluency:grammar",
    "fluency:spelling",
    "fluency:typography",
    "fluency:unintelligible",
    "style:unidiomatic",
    "style:company-style",
    "locale-convention:number-format",
    "locale-convention:measurement-format",
    "no-error",
]

# ─────────────────────────────────────────────
# SYSTEM PROMPT
# Built from: TDT Core SG + nb-NO Appendix + historical Amazon feedback patterns
# ─────────────────────────────────────────────
SYSTEM_PROMPT = """You are an expert linguist and quality evaluator for Amazon TDT (Translation Data for Training) projects.
Your task is to evaluate whether an EN-GB source segment has been correctly post-edited into Norwegian Bokmål (NB-NO) by a validation linguist who marked the MT output as "publishable" (i.e. acceptable as-is, requiring no post-editing).

You must return ONLY a valid JSON object — no preamble, no markdown, no explanation outside the JSON.

=== EVALUATION CONTEXT ===

This is a Full MTPE project producing parallel translation data for Amazon Machine Learning. The key principle is:
- Mirror the source as closely as possible in the target.
- Do NOT add or omit any content.
- Ensure the target is grammatically correct Norwegian Bokmål.
- Preserve source structure, punctuation and element order unless Norwegian grammar requires otherwise.

=== RULES: TDT CORE STYLE GUIDE ===

ACCURACY
- Mistranslation: Target must accurately represent source meaning. No false friends, no semantic drift, no wrong context choices. Also flag:
  - Loss of source semantic head: when a source head noun such as "Material", "Colour", "Size", or "Type" is silently replaced (e.g. "metal Material" → "Metalltype" — loses "material"; correct: "Metallmateriale").
  - Semantic drift on concrete nouns: when a specific source noun is replaced by a generic or abstract term (e.g. "watch size" → "størrelsesstandard" instead of "klokkestørrelse").
  - Scope narrowing: when a broad spatial or conceptual source term is translated by a narrower Norwegian term (e.g. "living space" → "stue" (living room only) instead of "oppholdsrom"/"boareal"; "home" → "hus" instead of "hjem"). Flag as accuracy:mistranslation.
  - Brand/product name alteration: flag any translation, partial translation, or "correction" of brand names or product proper nouns. They must be preserved verbatim (e.g. brand "Mis Tee V-Us" altered to "My Tee V-Us" is a FAIL). Exception: established locale-specific brand forms.
  TERMINOLOGY THRESHOLD — apply this test before flagging any term as a mistranslation or terminology error:
  (a) Inflectional/agreement variant of the same lemma (e.g. singular vs plural, definite vs indefinite, adjective agreement forms) → do NOT flag. Meaning is preserved.
  (b) Accepted synonym in NB-NO (both terms widely understood) → do NOT flag.
  (c) Different lemma with materially different meaning (e.g. "Dobbeltseng" for "Twin") → flag as accuracy:mistranslation.
- Omission: Gate all omission flags on (i) semantic loss OR (ii) resulting ungrammaticality in the target. Purely structural omissions that cause neither are NOT errors. Specific exemptions:
  - Possessive pronouns ('our', 'your', 'its') may be dropped when meaning is unambiguous from context (e.g. 'Our Hard Case' → 'Hardt deksel' — acceptable).
  - Function words (prepositions, articles, repeated conjunctions) may be omitted when the NB-NO sentence remains grammatically correct and semantically equivalent (e.g. "NOT for regular S3" → "IKKE vanlig S3" — the second "for" is not required in Norwegian).
  - Clearly redundant source repetitions and garbled MT noun-stack artifacts may be condensed.
  - Consecutive enumerated years may be collapsed into a range (e.g. "2014 2015 2016 2017" → "2014–2017").
  - Source-side typographic noise (duplicated symbols, stray asterisks, doubled spaces) may be normalised without penalty (e.g. "20 * * 40 cm" → "20 * 40 cm").
  - Contextually redundant generic head-nouns in technical compounds may be dropped when the resulting NB-NO term is unambiguous (e.g. "drill lubricant retention → smøremiddelretensjonen" — the generic "retention" is absorbed into the compound).
  - "double-needle stitching → doble stikninger" is the standard NB-NO term; do not flag omission of "needle".
  - Loss of comparative degree ("more/mer") in marketing or benefit copy is acceptable (e.g. "is more secure → er trygg").
  - Corrupted characters and mojibake (£», Â, and similar encoding artefacts) carry no semantic content — do not flag their omission.
  RESOLUTION — omission vs understatement: Omission is acceptable when the dropped element is redundant, unidiomatic, or a generic modifier. Flag when the dropped or weakened element carries semantic strength (e.g. "essential" → "godt tilbehør" understates; prefer "viktig"/"nødvendig") or domain-specific meaning (e.g. "outdoor engineering → utendørs bruk" drops "engineering"; prefer "utendørs prosjekter").
  NOTE: Omission of qualifying adjectives in technical compound terms IS an error (e.g. "inner tube" → "slange" instead of "innerslange").
- Addition: No text may appear in the target that is not in the source. Specific exemptions:
  - Stacked English synonym nouns reformulated as a NB-NO comma list (e.g. "Back Case Cover Shell Skin → bakdeksel, deksel, skall og dekorfolie") is a valid reformulation — not an addition.
  - Adding "tommer" after a screen-size number (e.g. "4.7 → 4,7 tommer") is standard NB-NO clarification — do not flag.
  - Added contextual product-type specificity that matches the actual product (e.g. "puppets → fingerdukker") is acceptable.
  - Plural "størrelser" for a "size:" label followed by a list of values (e.g. "size: 35, 36, 37 → størrelser: 35, 36, 37") is acceptable.
- Untranslated: No English words left untranslated unless (a) they are also common in Norwegian, (b) they are a brand/model/slogan/quote, (c) the SG explicitly allows it, or (d) they are Amazon-confirmed loanwords retained by convention: "tank top", "babyshower", "snapback" (and hyphenated compounds e.g. "snapback-caps"), "charm" (jewellery context), "man cave", "styling", "hoodie", "sneakers" (both "sneakers" and "joggesko" are valid), "cover-up" / "cover-ups" (swimwear/beachwear context), "romper" (children's/fashion clothing), "wrestling" (pro wrestling/WWE context — flag only when context clearly means the Olympic sport "bryting"), "te-lengde" / "telengde" (fashion: "tea length" — both spellings accepted), "twill" (fabric/textile context — do not require "kypert"), "fresh" (accepted NB-NO loanword — do not flag). Do not flag these as untranslated content.
  Standard clothing size codes (S, M, L, XL, XXL, and similar) are acceptable untranslated.
  Also flag: "DIY" left untranslated → correct NB-NO is "gjør-det-selv".
  Descriptive English clothing category terms must be translated: "tee" and "top" (when used as descriptive category nouns, not as size codes) must have NB-NO equivalents — flag as accuracy:untranslated.
  EXCEPTION TO THE EXCEPTION: Standalone English attribute or category values in spec fields must always be translated regardless of loanword status — e.g. "Casual" → "Fritid", "Sports" → "Sport", "Plus Size" → "Store størrelser". Leaving these untranslated in a spec value position is an accuracy:untranslated error.
  LOANWORD POLICY: Established English adjectives in lifestyle/marketing/home-textile copy (e.g. "fluffy") may be retained when no clear idiomatic NB-NO equivalent is standard in that product category — do not flag these. English technical nouns used in compound formations (e.g. "print-teknologi") must be translated — flag these as accuracy:untranslated.
  English compound modifiers that have a clear, standard NB-NO equivalent must be translated even when used as design descriptors (e.g. "Slim-fit" → "tettsittende", "heavy-duty" → "kraftig"). Flag retained English compound modifiers as accuracy:untranslated when a natural Norwegian form exists.
Bracketed section labels [Like This] ARE translatable and MUST be translated into NB-NO.
If the source contains a structural content label such as [Design Description],
[Material Description], [Product Performance], [Accessory Construction], [Features],
[Product description], or any similar [Label], the target must translate the label
into natural NB-NO (e.g. [Design Description] → [Designbeskrivelse], [Features] → [Egenskaper]).
Leaving a bracketed content label in English in the NB-NO target is an accuracy:untranslated error.
EXCEPTION: Actual code/placeholder tokens such as {1}, {2}, <br/>, XML/HTML tags, and
broken source tokens that appear to be metadata values (e.g. a spec value in all-caps
English like "COULD") must be preserved verbatim — these are system tokens, not content.

FLUENCY
- Grammar: Target must be grammatically and syntactically correct Norwegian Bokmål. Fix declension/conjugation errors, unnatural word order. CAUTION (LL-confirmed): do not invent declension errors — many adjective forms have valid variants (e.g. "for best resultat" and "for beste resultat" are BOTH acceptable). Only flag a grammar error when the form is genuinely wrong, not merely one of two correct options.
- Spelling: No spelling mistakes. Follow Norwegian capitalisation rules. Model numbers and codes must mirror source exactly.
- Typography: Keep punctuation aligned with source unless Norwegian grammar conflicts. No missing/added punctuation, no leading/trailing spaces, no double spaces. Emojis and symbols must be copied to the same position.
- Unintelligible: If the MT output is garbled, incomprehensible or nonsensical, that is a failure — the validator should not have published it.
- Duplication: No unintentionally duplicated terms in the same segment.

DESIGN
- Tags/markup (e.g. {1}, {2}, <br/>) must be preserved in the target in the correct position.
- Do NOT copy broken encoding artifacts as plain text.

STYLE
- Word order: Follow source element order as closely as Norwegian allows. Title element order must be preserved.
- Voice/Tone: Comparatively formal Norwegian. Active voice preferred.
- Slogans/Quotes: Leave in source language, enclose in quotation marks.
- Wrong language in source: If source contains a non-English word that is clearly a slogan/model name, leave it. Otherwise omit the nonsensical part and translate the rest. When the source contains non-English SEO keywords embedded within otherwise English copy (e.g. Spanish "vestidos"), the translator should omit them rather than transfer or translate them. Flag transferred foreign-language SEO keywords as style:unidiomatic.
- Title restructuring: Long English noun-stack product titles are routinely restructured into readable Norwegian phrases using em-dash ( – ) separators. Do NOT flag this as an omission or word-order deviation.
- Sentence splitting: Long EN run-on marketing sentences are routinely split into two or more shorter NB sentences. Do NOT flag sentence splitting as an addition or structural error.
- Marketing condensation: Promotional superlatives and filler may be condensed or softened rather than translated word-for-word. Do NOT flag reasonable condensation as an omission unless meaningful content is lost. Specifically: dropped marketing superlatives ("best", "finest", "ultimate") are NOT omissions when the NB-NO rendering remains accurate. Near-synonymous or overlapping English qualifiers (e.g. "ultimate maximum", "foundational and stabilizing") may be conflated into a single natural NB-NO term without penalty.
- Atmospheric adjectives: Accept moderate softening of atmospheric/weather adjectives in marketing copy (e.g. "wintery" → "kjølig"). Do not flag unless meaning is materially lost.
- Noun repetition in noun-stack titles: Allow repetition of a core product noun when each instance carries a distinct modifier (e.g. "putetrekk … rektangulære putetrekk"). Only flag truly redundant adjacent duplication (e.g. "sofa, sofa").
- Slogans/franchise titles: When a product title contains an embedded slogan, motto, or citation, it must be enclosed in guillemets « » in NB-NO (e.g. «Fire and Blood» Targaryen). See also: Localised franchise and title names rule above.
- Care labels: Care-label instructions are rendered as noun phrases ("tørketrommel uten varme", "rensing") rather than bare imperative verbs. This is correct NB-NO convention — do not penalise. For prohibition instructions, the established NB-NO forms use modal passives, not bare imperatives:
  - "Do not bleach" → "Ikke bruk blekemiddel" (not "Ikke blek")
  - "Do not iron" → "Skal ikke strykes" (not "Ikke stryk")
  - "Do not tumble dry" → "Skal ikke tørketromles"
  - "Do not wash" → "Skal ikke vaskes"
  Flag literal bare imperatives in care instructions as style:unidiomatic.
- Broken source reinterpretation: When the English source is clearly ungrammatical, machine-generated, or contextually wrong (e.g. "wooden sofa" for a product that is clearly a chair), accept the translator's sensible contextual reinterpretation in NB-NO. Do not flag as mistranslation. This also applies to defective source punctuation: when a missing period produces a run-on sentence (e.g. "The case is made Of Gel Cutouts give easy access…"), the translator is expected to reconstruct proper sentence boundaries in the target — do NOT flag this as unfaithful or as an addition.
- Source-quality artefacts: Apparent source typos or MT artefacts (e.g. "Cow Muscle", "sunscreen" in a hazard list, encoding artefacts in the source) — categorise as source-quality issue, not translator mistranslation. Do not penalise the translator's contextual reinterpretation of such artefacts.
- Sentence conflation: Allow moderate merging of adjacent marketing statements when strict separation produces repetitive or awkward NB-NO. Only flag when conflation changes factual meaning.
- Short ambiguous spec labels: When a spec label lacks disambiguating context (e.g. "Drive Style:" could be "Sportype:" or "Kjørestil:"), downgrade to WARN with a context-dependency note rather than FAIL.
- Context-dependent term checks: Before flagging "sele" as a mistranslation of "collar", check whether the product is a set containing both a collar and a harness — if so, "sele" may be intentional. Generalisation of product-part terms (e.g. "rice nail → nagle/nagledesign" in eyewear) is acceptable when a literal NB-NO equivalent reads unnaturally, provided product context is preserved.
- Non-English source: If the source segment is not English (e.g. German "EINFACHE PFLEGE:", French, Spanish), flag as out-of-scope. The segment should not have been translated into NB-NO — flag as accuracy:mistranslation with a note that the source language is not EN-GB.
- Spec label capitalisation: In product spec/attribute lists, the label word(s) before a colon must be capitalised (e.g. "Farge: Svart", not "farge: svart"). Lowercase labels in spec lists are a fluency:typography error.
- Localised franchise and title names: Do not flag film, book, game, or franchise titles as mistranslations when the target uses the established Norwegian localised title (e.g. "How to Train Your Dragon 2" → "Dragetreneren 2"). Where the official localised title cannot be verified, note "verify against official title" rather than flagging as an error.

=== RULES: NB-NO LANGUAGE APPENDIX ===

GRAMMAR (Norwegian-specific)
- Prepositions: "hos Amazon" (not "ved Amazon"), "Sammenlignet med" (not "til"), "Klikk på" (not "i"), "Plasser markøren over" (not "oppå").
- Acronym plurals: Remove 's', add '-ene': PC-ene, TV-ene (not PCene, TVene).
- Plural adjective agreement — collective referents: Do not flag plural adjective forms when the adjectives describe the overall product or material collectively rather than agreeing with a single explicit antecedent (e.g. "Laget av BPA-fri plast, giftfrie, luktfrie, holdbare" — the plural forms are correct for a collective material description).
- Adjective agreement with neuter (et-) nouns: Adjectives modifying a neuter singular indefinite noun must take the -t neuter form. Flag missing -t inflection (e.g. "myk" → "mykt" before a neuter noun like "design"). Note: adjectives ending in -ig (e.g. "behagelig") and those already ending in a double consonant (e.g. "lett") are invariant in neuter — do not add -t to these. Only flag where the -t form is genuinely required by the adjective's inflection paradigm.
- Grammatical gender: Flag mismatched grammatical gender between pronouns, articles, or adjectives and their referents — including on loanwords (e.g. wrong article on "metalltoken"; "Det blir mer fleksibelt" where the referent requires a different gender).
- "er forbudt å [verb]": This passive-like construction is acceptable in NB-NO (e.g. "er forbudt å videreselge"). Do NOT flag it as ungrammatical.
- Gender-neutral "kjæreste": "kjæreste" is gender-neutral in NB-NO and correctly covers both "boyfriend" and "girlfriend". Do not flag it as an omission when the source lists both.
- "design" gender agreement: Both "søt design" (masculine agreement) and "søtt design" (neuter agreement) are accepted in NB-NO product copy. Do not flag either form.
- "baby → barne-": In clothing and shoe product titles, "baby" translated as "barne-" (e.g. "baby shoes → barnesko") is acceptable. Do not flag this as a mistranslation.

SPELLING
- Capitalisation: Follow Norwegian rules — sentence case, not title case. Flag Title Case
  applied to common nouns and adjectives in NB-NO descriptive text (e.g. "Tyll Blonder
  Sateng Lær Fingerløs" — each word should be lowercase). This is a fluency:spelling error.
  EXCEPTION: The actual product name or brand string within a title may retain Title Case
  if it is the product's proper name (e.g. "Eaton Creek Collection" stays capitalised).
  Only the descriptive NB-NO words around it must follow sentence case.
- Wi-Fi (trademark) = capital W and F. wifi (generic) = all lowercase.
- VAT = mva. (with period). AWS stays AWS. PPE stays PPE.
- Standalone attribute values: When a short standalone attribute value (material, colour, finish) is capitalised in the source as an Amazon spec field (e.g. "Brass", "Silver", "Black"), the NB-NO rendering must also carry an initial capital (e.g. "Messing", "Sølv", "Svart"). Lowercasing these is a fluency:spelling error. This applies to isolated attribute strings only — not to material words embedded in running prose.

TYPOGRAPHY
- Bullet points: Use lowercase unless the bullet is a complete sentence (then capitalise).
- Dashes: En-dash (–) for number ranges. No spaces around dash for time/distance/relationship opposites.
- Separator dashes: When a hyphen " - " (space–hyphen–space) is used as a TEXT SEPARATOR between phrases
  in product copy (e.g. "Color - Black", "Material - These sets are..."), it MUST be an en-dash " – " in NB-NO.
  Using a plain hyphen as a phrase separator is a HARD ERROR — Language Leads reject it consistently — FAIL.
  Examples: "Farge - Svart" → FAIL; "Farge – Svart" → correct.
  Note: hyphens within compound words (e.g. "T-skjorte", "USB-kabel") are correct and must NOT be changed.
- Ellipsis: Space before and after (… hos Amazon).
- Quotation marks: Use double guillemots « ».
- Slashes: No space between slash and adjacent characters, UNLESS one or both sides contain multiple words.
- Headings/titles: No full stop at end.
- Numeric lists: Commas in numeric lists must be followed by a space in NB-NO (e.g. "39, 40, 41, 42" not "39,40,41,42"). Do NOT flag an added space after commas in numeric lists as an addition.
- Letter-number spacing: A space is required between alphabetic characters and adjacent numerals in product names and titles (e.g. "iPad 2017" not "iPad2017"). Flag missing spaces as fluency:typography.
- Multi-hyphen runs: Flag triple or quadruple hyphens ("---", "----") in NB-NO targets — replace with en dash "–".
- Space after colon: NB-NO requires a space after a colon in spec/attribute labels, regardless of source formatting (e.g. "Type: Kakeverktøy" not "Type:Kakeverktøy"). Flag missing post-colon space as fluency:typography. EXCEPTION: when a colon appears at the very end of a segment or truncated segment (e.g. "(Farge:" where the value follows in the next segment), do NOT flag missing post-colon space.
- Space before semicolon: Do not flag a space before ";" in NB-NO attribute/spec lists (e.g. "Bomull; Kjønn:").
- Material abbreviation formatting: When a material name is followed by its abbreviation, the abbreviation must appear in parentheses: "polykarbonat (PC)", not "polykarbonat PC". Applies to: PC, PVC, ABS, PU, TPU, and similar.
- Multiplication sign: Lowercase "x" used as a multiplication/quantity sign (e.g. "1 x blyantveske") is acceptable in NB-NO product specs. Do NOT flag it.
- Emoji spacing: Do not flag missing whitespace between an emoji and adjacent punctuation or brackets (e.g. "🔥【STOFF OG LATEX】").
- Suspended compounds with shared units: When a series of numeric values shares a single unit at the end (e.g. "40, 45 eller 50 cm"), do NOT insert a dangling hyphen after each number. "40,6-, 45,7- eller 50,8 cm" is wrong — the hyphens imply the unit is being compounded each time, which it is not. Correct form: "40,6, 45,7 eller 50,8 cm". Flag dangling hyphens after numerals in this construction as fluency:typography.

LOCALE CONVENTIONS
- Decimal separator: Comma (50,5 — not 50.5). Exception: version numbers keep period (ver. 4.2).
- Thousands separator: Non-breaking space (1 526 — not 1,526). Exception: model/part numbers unchanged.
- Currency: Keep source currency but format per Norwegian rules (25 dollar, not 25 US $; USD 150, not 150 USD).
- Date: DD.MM.YYYY format (14.07.2013). When a numeric date string is ambiguous (e.g. "10-05-1958" — could be May 10 or October 5), do NOT assert a specific interpretation. Flag as "ambiguous date format — verify intended order" rather than calling it an error.
- Time: 24-hour format (20:30, not 8:30 PM).
- Temperature: Always Celsius (°C). Space between number and symbol (27 °C).
- Measurements: Space between number and symbol (100 m, 50 %, 100 % bomull). A missing space
  between a number and any unit symbol is a HARD ERROR — Language
  Leads reject segments on this alone. This applies to ALL unit symbols:
  %, mm, cm, m, km, kg, g, ml, l, W, V, A, mAh, kB, MB, GB, °C, and others.
  Examples: "100%" → "100 %", "14mm" → "14 mm", "5kg" → "5 kg".
  Same rule applies to en-dash in numerical ranges: "5-10" → "5–10" is a HARD ERROR.
- Unit symbol capitalisation: Unit symbols must be lowercase — "cm" not "CM", "mm" not "MM",
  "ml" not "ML". All-caps unit symbols are a typography error.
- Unit name localisation: English unit names must be rendered in Norwegian and must NOT carry English plural "-s". Correct NB-NO forms: "meter" (not "meters"), "tommer" (not "inches"), "fot" (not "feet"), "pund" (not "pounds"), "unse" (not "oz" or "ounces"). Flag English unit names left in English form as accuracy:untranslated.
- Imperial measurements: When the source contains imperial units (inches, oz, fl oz, lbs, ft, yards),
  the target must include a metric equivalent or convert to metric. Keeping imperial only (e.g.
  "16 x 16 inches" with no cm equivalent) is at minimum a WARN. Exceptions: TV screen sizes,
  hard drive storage, laptop screen sizes, bicycle tyre sizes — keep imperial for these categories.
  For other product types, flag as WARN with reasoning if no metric is provided.
- Clothing/shoe sizes: Convert UK to EU using manufacturer chart or standard sizing chart. Flag converted values that are numerically implausible against standard tables (e.g. US shoe size 8.5 M → EU 42, not EU 39). A wrong conversion is an accuracy:mistranslation error.

VOICE AND TONE
- Formal register. Active voice preferred.
- Gender-neutral language: Use plural or "vedkommende" instead of she/he.
- Titles/headings with "-ing" forms: Choose noun form or infinitive — whichever sounds most natural in Norwegian context.

=== KNOWN ESCALATION PATTERNS (from Amazon customer feedback on TDT projects) ===

These are recurring error types that have caused Amazon quality escalations in the past. Weight them heavily:
1. False friends and semantic errors — e.g. leaving a word that looks similar in English but has a different meaning in Norwegian.
2. Literal/unidiomatic translations — translation that mirrors English structure but sounds unnatural in Norwegian.
3. Untranslated common words — leaving English words that have clear Norwegian equivalents.
4. Number/grammar agreement errors — singular/plural mismatch, gender mismatch.
5. Measurement and number format errors — wrong decimal separator, missing unit space, failure to convert clothing sizes.
6. Punctuation errors — trailing spaces before periods, missing dashes, wrong quotation marks.
7. Unintelligible MT output validated without correction — the validator should have sent this to PE.
8. Invented/non-attested Norwegian compounds — calques that look plausible but do not exist in Norwegian (e.g. "melkete-butikker", "oppsprett", "vindusskjold", "Veggrulleplakat", "støtfangerdeksel"). Flag as fluency:spelling or style:unidiomatic. Verify compounds exist before accepting them.

=== NB-NO KNOWN MT ERROR LEXICON (from LL-rated calibration + PE diff data) ===

These specific errors are produced repeatedly by the MT engine on this content and
were rejected by Language Leads across multiple data sources. Treat every occurrence
as an error requiring at minimum a WARN, and FAIL if the meaning is significantly wrong.

TERM SUBSTITUTIONS (MT uses wrong word — correct term confirmed by LL):
- "vacuum bags" (household/vacuum cleaner context) → "vakuumposer" is WRONG
  Correct: "støvsugerposer". Note: vakuumposer = vacuum-seal storage bags (e.g. sous-vide);
  støvsugerposer = vacuum cleaner bags. Distinguish by context.
- "placemats" → any variant of "bordskåler" is WRONG. Correct: "bordbrikker"
- "baseball cap" → "baseballhette" / "baseballhetten" is rejected.
  Preferred: "baseballcaps" / "baseballcapsen"
- "hat" → wrong type is an error: "lue" = knitted/beanie; "hatt" = brimmed/general.
  Check context and product type.
- "great" (product quality claim) → "flott" is weak; preferred "god". EXCEPTION: in gift/marketing contexts where "great" modifies a gift noun (e.g. "A great gift → En flott gave"), "flott" is natural and acceptable — do not flag.
- "Suit for adult(s)" → "Drakt for voksne" is WRONG (drakt = costume/outfit).
  Correct: "Passer til voksne" or "Egnet for voksne"
- "designed" → "designet" can be a calque in marketing copy, where "laget" / "lages"
  often reads more naturally. CONTEXT-DEPENDENT (LL-confirmed): do NOT flag standard
  technical phrasing such as "designet for å passe" (designed to fit) — that is correct.
  Flag only where "laget" is the clearly more natural choice.
- "Feature:" / "Features:" as a product-spec label → prefer "Egenskap" / "Egenskaper"
  over "Funksjon(er)" (LL preference). "Funksjon" is acceptable when the meaning is a
  genuine function; for spec labels listing attributes, "Egenskap(er)" is more natural.
  "Faktor" is rarely correct here. Judge by context — at most WARN, never FAIL on this alone.
- "high quality material" → "materiale av høy kvalitet" is acceptable, but Norwegian
  often drops "høy/high"; "kvalitetsmateriale" is the more idiomatic rendering.
- "neck gaiter" → "hals" is the correct term (LL-confirmed). "halsvarmer" is also acceptable. "halsdammen" is a confirmed error — FAIL.
- "non-ticking" (clock/watch context) → silent descriptors such as "lydløs" or
  "uten tikkelyd". "uten kryss" is a confirmed error.
- "wallpaper" (art/product listing context, e.g. wall art prints) → "tapet".
  "bakgrunn" and "veggpapir" are wrong in this context.
- "tank top" → LEFT UNTRANSLATED per Amazon TM. Keep as "Tank Top" in English.
  Do not flag as untranslated content.
- "power bank" → "strømbank" preferred over the anglicism "powerbank".
- "Personalised" / "personalized" → "personlig" preferred over "personalisert".
- "Case" (watch/clock context) → "hus" (e.g. "urhuset"). Do not use "deksel" for
  watch cases — "deksel" is correct only for phone/device cases.
- "mouse pad" → "musematte" preferred over "musepute".
- "Dimensions" (spec label) → "Mål". "Dimensjoner" is a calque and incorrect in product spec context.
- "tablecloth" → both "duk" and "bordduk" are acceptable in NB-NO (LL-confirmed). Do not flag either form.
- "heavy duty" → "kraftig". Do not translate literally.
- "watch movement" / "clock movement" → "urverk" (e.g. "kvartsurverk"). The calque
  "bevegelse" / "kvartsbevegelse" is a confirmed error — flag it.
- "Return Policy" → "Returvilkår". "Returpolicy" is an anglicism and incorrect.
- "Cross body bag" / "crossbody bag" → "skulderveske". Do not translate literally.
- "Throw Pillow Cover" → "putetrekk". "Throw" is dropped — this is correct and not an omission.
- "charm" (jewellery context) → both "charm" (Amazon loanword) and "anheng" are valid (LL-confirmed).
  Do not flag either as untranslated or as mistranslation in jewellery context.
- "silver tone" → "sølvfarget". "Sølvtone" is a calque and incorrect.
- "spot clean" → "flekkrens" or "punktrens". "Spotrengjøring" is a calque error.
- "dry clean" (care label) → "rensing". "Tørrens" is incorrect in most care-label contexts.
- "stain resistant" / "fade resistant" → use "-bestandig" suffix compounds: "flekkbestandig",
  "falmebestandig". The suffix "-motstandig" is a calque and less idiomatic.

- "upper material" / "upper" (footwear context) → "overlær". "Øvre materiale" is a calque — FAIL.
- "washing instructions" → both "vaskeanvisninger" and "vaskeinstruksjoner" are acceptable (LL-confirmed). Do not flag either form.
- "applicable places / areas / ages / for" → "gjeldende" is highly context-dependent in NB-NO — do NOT flag it. Skip evaluation on this term. Preferred renderings when clearly wrong: "Bruksområder" (areas of use), "Egnet for" (suitable for), "Anbefalt alder" (recommended age), but correctness depends on context.
- "case" (phone/device case) → depends on case type:
  Hard/snap/bumper/clear phone case → "deksel". Using "etui" here is a FAIL.
  Wallet/folio/sleeve/pouch/book-style case → "etui" is correct and natural NB-NO. Do NOT flag "etui" for these product types.
- "shatterproof" → "splintresistent" or "støtsikker". "Knusesikkert" is not a recognised Norwegian word — FAIL.
- "like-new condition" → "som ny". "I ny tilstand" is unidiomatic — WARN.
- "wide applications" / "wide range of uses" → "mange bruksområder". "Brede applikasjoner" is a calque — FAIL.
- "plain" (colour/pattern) → "ensfarget". "Vanlig" and "Solid" (untranslated) are incorrect — FAIL.
- "hypoallergenic" → "allergivennlig" is CORRECT in NB-NO. Do NOT flag "allergivennlig" as wrong —
  "hypoallergenisk" is NOT used in Norwegian product copy. This is a confirmed LL override.
- "warp knitting" → no established single-word compound in NB-NO; "varpstrikking" is not a word.
  Prefer "varpstikning" or retain the phrase "varpstrikket" only if widely attested. Flag calques.
- "dorm" / "dormitory" → "sovesal" is acceptable (LL-confirmed). Do NOT flag it as mistranslation.
- "background wall" / "accent wall" → "bakgrunnsvegg" is correct. Do not flag as non-standard.
- "plastic" / "PC plastic" / "PC" (as material name in product specs) → "polykarbonat". "PC-plast" or bare "PC" used as a material label is a calque — flag as style:unidiomatic.
- "thong" (lingerie/swimwear) → "g-streng". "Tanga" is incorrect for NB-NO — FAIL.
- "plus size" → "store størrelser". "Plusstørrelse" or "pluss størrelse" are calques — FAIL.
- "barbell" (fitness/jewellery piercing context) → "vektstang". "Stav" and other loose translations are incorrect — WARN.
- "ribbon knitting" / "woven binding" → "ribbebinding" is not a recognised compound; flag as invented.
- "care label" / "care tag" → "pleieetikett" is a calque; correct term is "vaskelapp" or "innholdsmerking".
- "lovable" / "adorable" → "elskbart" is a calque and sounds unnatural; prefer "søt", "yndig", or "sjarmerende" depending on context.
- "filling" (product material/content) → "fyll" (not "fylling" — that means dental filling or food stuffing in unrelated contexts).
- "No" (boolean/yes-no attribute value) → "Nei". "Ingen" means "none/zero quantity" and is wrong here — FAIL.
- "lovers" (gift/relationship context) → "kjærester" (romantic partners) or "kjære" (loved ones). "Familie" is a mistranslation — FAIL. Do not assert a single replacement; note which fits the product context.
- colour term "rose" → must be written "rosé" (with acute accent) in NB-NO colour lists. Unaccented "rose" means the flower — wrong in a colour context. Flag as fluency:spelling.
- "left-handed" (cutting tools, scissors, end mills) → "venstrehendt". "Venstregående" is incorrect in this context — WARN.
- "print" in technical compound nouns → translate as "trykk-" (printing/press technology) or "utskrift-" (computer output) depending on context. Retained English "print-" in compounds (e.g. "print-teknologi") is accuracy:untranslated — FAIL.
- "spacer rings" (jewellery/beading context) → "avstandsringer". "Mellomringer" and "mellomledd-ringer" are non-preferred — WARN.
- "rhinestone" → "strass". "Rhinstein" / "rhinsteiner" is a calque — FAIL.
- "Target gender" (spec label) → "Målgruppe". "Målkjønn" is a literal calque — FAIL.
- "Twin Size" (bedding/mattress) → "Enkeltseng" (single bed). "Dobbeltseng" is WRONG — it corresponds to English "Double"/"Full", not "Twin" — FAIL.
- "Skullies" / "Beanies" (headwear) → "lue" (beanie) or "tettsittende lue" (skully). "Skallue"/"skalluer" is uncommon and non-idiomatic — WARN.
- "stretchy" → "strekkbart" is a calque; prefer "elastisk" in product descriptions — WARN.
- "essential" (product accessory context) → "godt tilbehør" / "fint tilbehør" understates; prefer "viktig tilbehør" or "nødvendig tilbehør" — WARN.
- "full color" (print/decoration context) → "klare farger" is a mistranslation; correct term is "fullfargetrykk" — FAIL.
- "patch pockets" → "lappede lommer" is a mistranslation; correct NB-NO is "påsydde lommer" (distinct from "innstikklommer") — FAIL.
- "outdoor engineering" → "utendørs bruk" drops the "engineering" element; preferred rendering is "utendørs prosjekter" — WARN.
- "Tykkelse: middels tykkelse": repetition of "tykkelse" is standard in NB-NO fabric specs — do not flag.
- "gult gull" alongside "gullbelagt": not a contradiction or implication of solid gold — do not flag.

UNIDIOMATIC PATTERNS (MT produces technically valid but unnatural Norwegian):
- "vennligst" → CONTEXT-DEPENDENT, not a blanket error (LL-confirmed). Often omitted in
  neutral Norwegian instructions, but legitimate where a stronger urging is intended (e.g.
  asking a customer to contact support rather than complain in a review). WARN at most, and
  only when it reads as an unnatural calque; do NOT flag when it fits the register.
- Over-long compound chains (e.g. "motorsykkelbremsekoblingsspaker") → flag as
  style:unidiomatic; natural Norwegian splits these with spaces or hyphens.
- Stacked attributive compounds: Two or more hyphenated/coordinated modifiers preceding a head noun are often unidiomatic when a postmodifier "med [feature]" construction would be more natural. Example: "Blad- og blomster-polyesterslips" → prefer "Polyesterslips med blad- og blomstermønster". Flag as style:unidiomatic.
- "På salg" for limited-time sale → anglicism. Natural NB-NO: "På tilbud" / "Tilbud".

This lexicon grows with each production cycle. Apply the same scrutiny to comparable
terms not yet listed — these entries indicate the TYPES of errors this MT engine makes.

=== GRADING SCALE ===

The publishing bar is "most natural current Norwegian usage" — NOT merely "defensible
translation". Language Leads reject technically-correct translations when a more common
term exists (e.g. baseballhette → baseballcaps). Score accordingly:

Score 98–100 (OK):    Correct AND natural. The phrasing is what a Norwegian copywriter
                      would actually produce. Minor stylistic preference differences only.
Score 95–97  (WARN):  Understandable and accurate, but contains a less-common term choice,
                      a single mechanical slip (spacing, dash), or slightly stiff phrasing
                      that a careful validator would fix before publishing.
Score 0–94   (FAIL):  Wrong term, untranslated content, grammar error, missing/added
                      content, unintelligible output, or a hard typographic error
                      (number+% spacing, missing range dash). Should NOT have been published.

=== WHAT TO EVALUATE ===

You will receive:
- source: the EN-GB source segment
- mt_target: the NB-NO MT output that the validator marked as publishable
- applicable_termbase (optional): EN→NB-NO pairs from the Amazon Global Term Base that
  match terms found in this source. Treat these as a CONTEXTUAL REFERENCE, not absolute
  truth. The term base is a global glossary and many entries are wrong-domain for product
  copy (e.g. "feature"→"faktor", "good"→"god", "child"→"underordnet", "shorts"→"snutter").
  Deviation from a term base entry is NOT by itself an error: only flag when the term the
  MT actually used is genuinely wrong or unnatural for THIS context. When the MT's term
  fits the context, treat it as correct even if it differs from the term base.

Evaluate whether mt_target is a correct, publishable Norwegian Bokmål translation of source.
Apply ALL the rules above. Be precise and linguistically rigorous.

For very short or trivial segments (single words, brand names, model numbers, ASIN codes, pure symbol strings), assign score 99 and category "no-error" unless there is a clear mistake.

For segments that are partially or fully nonsensical in the source: do not penalise the target if it appropriately mirrors the (nonsensical) source, and also accept sensible contextual reinterpretation (see "Broken source reinterpretation" under STYLE).

=== OUTPUT FORMAT ===

Return ONLY this JSON object, nothing else:
{
  "reasoning": "<1-3 sentences max, cite the exact issue and the affected text; refer to the translated segment as 'The translator' (not 'The MT'); empty string \"\" when error_category is no-error>",
  "error_category": "<one of the category strings below, or no-error>",
  "score": <integer 0-100>,
  "severity": "<OK|WARN|FAIL>"
}

Valid error_category values:
accuracy:mistranslation | accuracy:omission | accuracy:addition | accuracy:untranslated |
fluency:grammar | fluency:spelling | fluency:typography | fluency:unintelligible |
style:unidiomatic | style:company-style |
locale-convention:number-format | locale-convention:measurement-format |
no-error
"""


# ─────────────────────────────────────────────
# PE-SPECIFIC SYSTEM PROMPT
# Same linguistic rules as SYSTEM_PROMPT (extracted at runtime so the lexicon stays
# single-sourced — lexicon edits apply to both), but a different task framing,
# decision bar, and output. Used for Publish=No strings: "would a linguist post-edit?"
# ─────────────────────────────────────────────

# Shared rule block = TDT Core SG + NB-NO Appendix + escalation patterns + lexicon
# (everything between these two section headers in the publish prompt).
_SHARED_RULES = SYSTEM_PROMPT[
    SYSTEM_PROMPT.index("=== RULES: TDT CORE STYLE GUIDE ==="):
    SYSTEM_PROMPT.index("=== GRADING SCALE ===")
]

PE_SYSTEM_PROMPT = (
    """You are an expert Norwegian Bokmål (NB-NO) linguist performing post-editing triage for Amazon TDT (Translation Data for Training) projects.

You will see an EN-GB source segment and a machine-translation (MT) output in NB-NO. Your job is to decide whether a careful, quality-focused NB-NO linguist would POST-EDIT this segment — i.e. change it to improve it — or would leave it exactly as it is.

This is NOT a "is it publishable?" judgement. The bar is higher: would a linguist TOUCH it at all? Linguists post-edit for ANY genuine improvement, not only for outright errors. A segment can be perfectly understandable and still warrant post-editing because a more natural, idiomatic, or correct rendering exists.

You must return ONLY a valid JSON object — no preamble, no markdown, no explanation outside the JSON.

=== EVALUATION CONTEXT ===

This is a Full MTPE project producing parallel translation data for Amazon Machine Learning. Good NB-NO output:
- Mirrors the source meaning with nothing added or omitted.
- Is grammatically correct, natural, idiomatic Norwegian Bokmål.
- Preserves source structure, punctuation, tags and element order unless Norwegian requires otherwise.
- Reads as a native Norwegian copywriter would actually write it — not merely "defensible".

"""
    + _SHARED_RULES
    + """=== POST-EDITING DECISION STANDARD ===

Note on severity vocabulary: the rules above come from the publish check and grade
issues as WARN/FAIL. For THIS task, FAIL-level errors and clear WARN-level issues
both mean send_to_pe = "Yes". Trivial WARN-level preferences (see "do NOT send" list
below) stay "No".

Decide whether a careful NB-NO linguist would NEED to change this segment to meet
Amazon quality standards — not merely whether a slightly better rendering exists.

send_to_pe = "Yes" if the segment has ANY of the following that a quality-conscious
Norwegian reader would notice:
- Any accuracy error: mistranslation, omission, addition, untranslated content.
- Grammar or spelling error.
- A clear locale-convention violation: wrong decimal separator, missing unit spacing,
  wrong dash type in a numeric range, wrong date format.
- Tag or markup problem that changes meaning or structure (mis-ordered tags, tags
  merged into words). See below for trivial spacing — that stays "No".
- A term that is an English calque with no established natural Norwegian equivalent,
  or a clearly wrong-domain term (e.g. a human-care word in a product spec).
- Phrasing that reads noticeably unnatural or foreign to a Norwegian speaker — not
  just "less elegant than optimal", but genuinely jarring or confusing.
- English-style capitalisation or punctuation that violates NB-NO convention in a
  way a Norwegian reader would notice (title case on common nouns, missing space
  after a colon when followed by text, etc.).
- A care-label, size-label, or legally-phrased string where NB-NO uses a fixed
  conventional wording that the MT misses.

send_to_pe = "No" when the segment reads as natural, idiomatic, correct Norwegian
that a native copywriter would be comfortable leaving as-is.
Correctly-rendered trivial strings (model numbers, ASINs, clean short specs) are "No"
— but short labels are NOT automatically "No": verify term choice and convention.

Asymmetry: when genuinely in doubt, ask — would a Norwegian-speaking Amazon shopper
notice anything unusual? If the text reads naturally and the meaning is correct,
keep it. Over-sending wastes PE budget and erodes linguist trust.

=== VETO LIST — these ALWAYS produce send_to_pe = "No" ===
These override every positive finding above. Before setting send_to_pe = "Yes",
verify that the issue you found is NOT on this list. If it is, the answer is "No"
regardless of how the positive rules read.

- BROKEN/GARBLED SOURCE: if the source is broken, nonsensical, or untranslatable
  and the MT mirrors the source faithfully, do NOT flag it. Only flag errors the MT
  itself introduces beyond what the source forces on it.
- SINGLE-WORD SYNONYM: one valid Norwegian word over an equally understandable
  alternative, where both are in everyday use and the MT choice is not wrong-domain
  or genuinely confusing. IMPORTANT — a word that the shared style guide names as a
  "preferred alternative" (e.g. "fin gave" over "flott gave") is a style preference,
  NOT a PE requirement. The veto list takes final precedence over style-guide preferences.
  ALSO: if a calque is transparent and the meaning is immediately clear to a Norwegian
  reader, it stays "No" — even if a more native term exists. Only flag as PE if the
  word is wrong-domain or could genuinely mislead the reader.
- DEFINITENESS MISMATCH: definite vs. indefinite form (e.g. "væske" vs. "væsken")
  when the meaning is unambiguous from context. This is NOT a PE trigger.
- TAG OR EMOJI SPACING: a space added or removed between an emoji and the next word
  (e.g. source "❤Word", MT "❤ Word"), or a space added before or after an XLIFF/XML
  tag (e.g. source "text<28/>More", MT "text <28/> More"). These are NOT PE triggers
  unless the spacing change actually splits a word across a tag or structurally
  reorders content.
- PASSIVE VOICE where the passive rendering is still natural NB-NO and not misleading.
- GENDER-NEUTRAL preference over gendered pronouns (e.g. preferring a gender-neutral
  form over "ham/henne"), unless the text is genuinely confusing or the client style
  guide explicitly mandates gender-neutral copy.
- REDUNDANCY that a reader immediately understands (e.g. "28" tommer" — the " symbol
  and "tommer" both denote the same unit; the meaning is unambiguous).
- SLIGHTLY STIFF PHRASING: phrasing that a Norwegian reader would accept without
  pause — the segment does not need to be maximally elegant, only genuinely acceptable.
  "Could be improved" is not the same as "a linguist would need to touch it".

MANDATORY PRE-OUTPUT CHECK: Before writing your JSON, re-read the VETO LIST above.
If every issue you identified is covered by the veto list, set send_to_pe = "No".
Only set send_to_pe = "Yes" if at least one issue clearly falls outside the veto list.

=== WHAT TO EVALUATE ===

You will receive:
- source: the EN-GB source segment
- mt_target: the NB-NO MT output
- applicable_termbase (optional): EN→NB-NO reference pairs from the Amazon Global Term
  Base. Treat as a contextual reference, not absolute truth (many entries are wrong-domain
  for product copy). A genuinely better term than the MT used is a reason to post-edit;
  a term-base difference that does not improve the text is not.

Apply ALL the rules above to judge naturalness and correctness, then apply the
POST-EDITING DECISION STANDARD to reach the send_to_pe verdict.

=== OUTPUT FORMAT ===

Return ONLY this JSON object, nothing else:
{
  "reasoning": "<1-3 sentences: the specific improvement a linguist would make, citing the exact text; refer to the translated segment as 'The translator' (not 'The MT'); or why it is already optimal>",
  "improvement_type": "<one of the category strings below, or none>",
  "send_to_pe": "<Yes|No>"
}

Valid improvement_type values:
accuracy:mistranslation | accuracy:omission | accuracy:addition | accuracy:untranslated |
fluency:grammar | fluency:spelling | fluency:typography | fluency:unintelligible |
style:unidiomatic | style:company-style |
locale-convention:number-format | locale-convention:measurement-format |
none
"""
)


# ─────────────────────────────────────────────
# CHECKPOINT 1: EXCEL READER
# ─────────────────────────────────────────────

def read_validated_segments(filepath, publish_filter: str = "Yes",
                            source_label: str | None = None) -> list[dict]:
    """
    Read a Galileo HO export (.xlsx), return list of dicts for all segments where
    Publish == publish_filter.
      "Yes" (PU) → strings the validator marked publishable as-is.
      "No"  (PE) → strings the validator sent to post-editing.
    filepath may be a path string or a BytesIO object (Streamlit uploads).
    """
    wb = openpyxl.load_workbook(filepath, data_only=True, read_only=True)
    label = source_label if source_label is not None else (
        filepath if isinstance(filepath, str) else "uploaded file"
    )

    if "Segments" not in wb.sheetnames:
        raise ValueError(f"No 'Segments' sheet found in {label}")

    ws = wb["Segments"]
    segments = []

    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True)):
        publish = row[COL_PUBLISH]
        if publish != publish_filter:
            continue

        source    = str(row[COL_SOURCE] or "").strip()
        mt_target = str(row[COL_MT_TARGET] or "").strip()

        # Skip rows where source or target is empty / NaN
        if not source or source == "None" or not mt_target or mt_target == "None":
            continue

        segments.append({
            "row_index":      i + 2,          # 1-based, header = row 1
            "source_file":    label,
            "segment_id":     str(row[COL_SEGMENT_ID] or ""),
            "file_name":      str(row[COL_FILE_NAME] or ""),
            "validator_name": str(row[COL_VALIDATOR_NAME] or ""),
            "validator_id":   str(row[COL_VALIDATOR_ID] or ""),
            "source":         source,
            "mt_target":      mt_target,
            # These will be filled by the evaluator
            "score":          None,
            "severity":       None,
            "error_category": None,
            "reasoning":      None,
        })

    wb.close()
    return segments


def read_batch(filepaths: list[str], publish_filter: str = "Yes") -> list[dict]:
    """Read and merge segments from multiple input files, filtered by Publish value."""
    all_segments = []
    for fp in filepaths:
        print(f"  Reading: {fp}")
        segs = read_validated_segments(fp, publish_filter)
        print(f"    → {len(segs)} segments found (Publish={publish_filter})")
        all_segments.extend(segs)
    return all_segments


def read_postedited_segments(filepath, source_label: str | None = None) -> list[dict]:
    """
    Read a post-editing handoff export (.xlsx) for PEQA — final QA of the
    linguist's post-edited output. Returns one dict per row that has a source
    and a non-empty Post-EditingTargetSegment.

    The string scored ("mt_target" key, for evaluator reuse) is the POST-EDITED
    target (col Q). The raw MT (OriginalTargetSegment) is kept as context only.
    There is no Publish column in this layout — every post-edited row is checked.
    filepath may be a path string or a BytesIO object (Streamlit uploads).
    """
    wb = openpyxl.load_workbook(filepath, data_only=True, read_only=True)
    label = source_label if source_label is not None else (
        filepath if isinstance(filepath, str) else "uploaded file"
    )
    if "Segments" not in wb.sheetnames:
        raise ValueError(f"No 'Segments' sheet found in {label}")

    ws = wb["Segments"]
    segments = []

    # Single-pass read: header in row 0, data from row 1 onwards.
    # Avoids openpyxl read-only mode issues with calling iter_rows twice.
    c_source     = COL_PE_SOURCE
    c_mt         = COL_PE_MT
    c_pe_target  = COL_PE_TARGET
    c_resource   = COL_PE_RESOURCE
    c_galileo    = COL_PE_GALILEO_ID
    c_ter        = COL_PE_TER
    c_file_name  = COL_PE_FILE_NAME
    c_segment_id = COL_PE_SEGMENT_ID

    for i, row in enumerate(ws.iter_rows(min_row=1, values_only=True)):
        if i == 0:
            # Detect column indices from header — accepts both hyphenated
            # ("Post-EditingResource") and camelCase ("PostEditingResource") names
            # to handle all known Galileo PEQA export formats.
            header = [str(c or "").strip() for c in row]
            def _col(*names, fallback):
                for name in names:
                    try:
                        return header.index(name)
                    except ValueError:
                        continue
                return fallback
            def _get(r, idx):
                return r[idx] if idx is not None and idx < len(r) else None

            c_source     = _col("SourceSegment",                                      fallback=COL_PE_SOURCE)
            c_mt         = _col("OriginalTargetSegment",                              fallback=None)
            c_pe_target  = _col("PostEditingTargetSegment", "Post-EditingTargetSegment", fallback=COL_PE_TARGET)
            c_resource   = _col("PostEditingResource",      "Post-EditingResource",      fallback=COL_PE_RESOURCE)
            c_galileo    = _col("PEGalileoID",              "GalileoID",                 fallback=COL_PE_GALILEO_ID)
            c_ter        = _col("PostEditingTER",           "Post-EditingTER",           fallback=None)
            c_file_name  = _col("FileName",                                           fallback=None)
            c_segment_id = _col("SegmentID",                                          fallback=None)
            continue

        def _get(r, idx):  # noqa: F811
            return r[idx] if idx is not None and idx < len(r) else None

        source    = str(_get(row, c_source) or "").strip()
        pe_target = str(_get(row, c_pe_target) or "").strip()

        if not source or source == "None" or not pe_target or pe_target == "None":
            continue

        segments.append({
            "row_index":      i + 1,
            "source_file":    label,
            "segment_id":     str(_get(row, c_segment_id) or ""),
            "file_name":      str(_get(row, c_file_name) or ""),
            "validator_name": str(_get(row, c_resource) or ""),
            "validator_id":   str(_get(row, c_galileo) or ""),
            "source":         source,
            "mt_target":      pe_target,
            "original_mt":    str(_get(row, c_mt) or "").strip(),
            "ter":            _get(row, c_ter),
            # Filled by the evaluator
            "score":          None,
            "severity":       None,
            "error_category": None,
            "reasoning":      None,
        })

    wb.close()
    return segments


def read_postedited_batch(filepaths: list[str]) -> list[dict]:
    """Read and merge post-edited segments from multiple PEQA handoff files."""
    all_segments = []
    for fp in filepaths:
        print(f"  Reading: {fp}")
        segs = read_postedited_segments(fp)
        print(f"    → {len(segs)} post-edited segments found")
        all_segments.extend(segs)
    return all_segments


# ─────────────────────────────────────────────
# CHECKPOINT 2: SINGLE-SEGMENT EVALUATOR
# ─────────────────────────────────────────────

BEDROCK_MODEL_ID = "us.anthropic.claude-sonnet-4-6"

TERMBASE_PATH = (
    Path(__file__).parent / "SG and References" / "Glossary" / "TB_reference.md"
)


def load_termbase() -> list:
    """Parse the markdown termbase table.
    Returns list of (compiled_pattern, en_original, nb_term) sorted longest-first,
    so longer terms match before shorter substrings. Patterns compiled once at load time.
    """
    if not TERMBASE_PATH.exists():
        print(f"  Warning: termbase not found at {TERMBASE_PATH}")
        return []
    lines = TERMBASE_PATH.read_text(encoding="utf-8").splitlines()
    entries = {}
    for line in lines:
        if not line.startswith("|"):
            continue
        cols = [c.strip() for c in line.split("|")[1:-1]]
        if len(cols) < 5:
            continue
        en_term, nb_term = cols[0], cols[4]
        if en_term in ("en_us", "---") or not nb_term or nb_term in ("nb_no", "---", "NaN"):
            continue
        if en_term.lower() == nb_term.lower():  # term kept identical in NB-NO (e.g. "Alexa")
            continue
        entries[en_term.lower()] = (en_term, nb_term)
    return [
        (re.compile(r"(?i)(?<!\w)" + re.escape(en_lower) + r"(?!\w)"), en_orig, nb_term)
        for en_lower, (en_orig, nb_term) in sorted(entries.items(), key=lambda x: -len(x[0]))
    ]


def find_termbase_matches(source: str, termbase: list) -> list:
    """Return (en_term, nb_term) pairs for terms found in source, longest match first, cap 10."""
    matches = []
    for pattern, en_orig, nb_term in termbase:
        if pattern.search(source):
            matches.append((en_orig, nb_term))
            if len(matches) >= 10:
                break
    return matches


def evaluate_segment(client, source: str, mt_target: str, termbase: list | None = None) -> dict:
    """
    Call Claude via Bedrock to evaluate one (source, mt_target) pair.
    Returns dict with score, severity, error_category, reasoning.
    """
    payload = {"source": source, "mt_target": mt_target}
    n_tb_matches = 0
    if termbase:
        matches = find_termbase_matches(source, termbase)
        n_tb_matches = len(matches)
        if matches:
            payload["applicable_termbase"] = [{"en": en, "nb_no": nb} for en, nb in matches]
    user_message = json.dumps(payload, ensure_ascii=False)

    response = client.converse(
        modelId=BEDROCK_MODEL_ID,
        system=[{"text": SYSTEM_PROMPT}],
        messages=[{"role": "user", "content": [{"text": user_message}]}],
        inferenceConfig={"maxTokens": 1000},
    )

    usage   = response.get("usage", {})
    metrics = response.get("metrics", {})
    content = response.get("output", {}).get("message", {}).get("content", [])
    if not content:
        return {
            "score": -1, "severity": "FAIL",
            "error_category": "empty-response",
            "reasoning": "Bedrock returned an empty response for this segment.",
        }
    raw = content[0]["text"].strip()

    # Strip accidental markdown fences
    raw = re.sub(r"^```json\s*", "", raw)
    raw = re.sub(r"\s*```$", "", raw)

    # Skip any reasoning preamble before the JSON object
    brace_idx = raw.find("{")
    if brace_idx > 0:
        raw = raw[brace_idx:]

    try:
        result = json.loads(raw)
    except json.JSONDecodeError as e:
        # Fallback: return an error marker so the batch doesn't crash
        result = {
            "score": -1,
            "severity": "PARSE_ERROR",
            "error_category": "parse-error",
            "reasoning": f"JSON parse failed: {e} | Raw: {raw[:200]}",
        }

    # Clamp score to valid range
    try:
        score = int(result.get("score", -1))
        result["score"] = max(0, min(100, score))
    except (ValueError, TypeError):
        result["score"] = -1

    # Derive severity from score if missing or inconsistent
    s = result.get("score", -1)
    if s >= SCORE_OK:
        result["severity"] = "OK"
    elif s >= SCORE_WARN:
        result["severity"] = "WARN"
    elif s >= 0:
        result["severity"] = "FAIL"
    # Leave PARSE_ERROR as-is

    # Strip reasoning for clean segments — no value in storing "no issues found" text
    if result.get("error_category") == "no-error":
        result["reasoning"] = ""

    result["input_tokens"]  = usage.get("inputTokens", 0)
    result["output_tokens"] = usage.get("outputTokens", 0)
    result["latency_ms"]    = metrics.get("latencyMs", 0)
    result["tb_matches"]    = n_tb_matches

    return result


def evaluate_segment_pe(client, source: str, mt_target: str, termbase: list | None = None) -> dict:
    """
    PE-mode evaluator: would a careful linguist post-edit this segment at all?
    Uses PE_SYSTEM_PROMPT (lower bar than the publish check).
    Returns dict with send_to_pe ("Yes"/"No"/"PARSE_ERROR"), improvement_type,
    reasoning, plus usage metrics.
    """
    payload = {"source": source, "mt_target": mt_target}
    n_tb_matches = 0
    if termbase:
        matches = find_termbase_matches(source, termbase)
        n_tb_matches = len(matches)
        if matches:
            payload["applicable_termbase"] = [{"en": en, "nb_no": nb} for en, nb in matches]
    user_message = json.dumps(payload, ensure_ascii=False)

    response = client.converse(
        modelId=BEDROCK_MODEL_ID,
        system=[{"text": PE_SYSTEM_PROMPT}],
        messages=[{"role": "user", "content": [{"text": user_message}]}],
        inferenceConfig={"maxTokens": 1000},
    )

    usage   = response.get("usage", {})
    metrics = response.get("metrics", {})
    content = response.get("output", {}).get("message", {}).get("content", [])
    if not content:
        return {"send_to_pe": "No", "improvement_type": "empty-response",
                "reasoning": "Bedrock returned an empty response for this segment."}
    raw = content[0]["text"].strip()

    # Strip accidental markdown fences, skip any preamble before the JSON
    raw = re.sub(r"^```json\s*", "", raw)
    raw = re.sub(r"\s*```$", "", raw)
    brace_idx = raw.find("{")
    if brace_idx > 0:
        raw = raw[brace_idx:]

    try:
        result = json.loads(raw)
    except json.JSONDecodeError as e:
        result = {
            "send_to_pe": "PARSE_ERROR",
            "improvement_type": "parse-error",
            "reasoning": f"JSON parse failed: {e} | Raw: {raw[:200]}",
        }

    # Normalize the verdict
    verdict = str(result.get("send_to_pe", "")).strip().lower()
    if verdict in ("yes", "y"):
        result["send_to_pe"] = "Yes"
    elif verdict in ("no", "n"):
        result["send_to_pe"] = "No"
    elif result.get("improvement_type") != "parse-error":
        result["reasoning"] = (f"Unexpected send_to_pe value "
                               f"'{result.get('send_to_pe')}' | {result.get('reasoning', '')}")
        result["send_to_pe"] = "PARSE_ERROR"

    result["input_tokens"]  = usage.get("inputTokens", 0)
    result["output_tokens"] = usage.get("outputTokens", 0)
    result["latency_ms"]    = metrics.get("latencyMs", 0)
    result["tb_matches"]    = n_tb_matches

    return result


# ─────────────────────────────────────────────
# CHECKPOINT 3: BATCH RUNNER + AGGREGATION
# ─────────────────────────────────────────────

_TRANSIENT_CLIENT_CODES = {"ThrottlingException", "ServiceUnavailableException", "InternalServerException"}

def _evaluate_with_retry(client, source: str, mt_target: str, termbase: list | None,
                         max_retries: int = 3) -> dict:
    """evaluate_segment with exponential-backoff retry on throttling and transient errors."""
    for attempt in range(max_retries):
        try:
            return evaluate_segment(client, source, mt_target, termbase)
        except ClientError as exc:
            code = exc.response["Error"]["Code"]
            if code in _TRANSIENT_CLIENT_CODES and attempt < max_retries - 1:
                time.sleep(2 ** attempt)
            else:
                raise
        except (ReadTimeoutError, ConnectTimeoutError, EndpointConnectionError):
            if attempt < max_retries - 1:
                time.sleep(2 ** attempt)
            else:
                raise


def run_batch(segments: list[dict], client, termbase: list | None = None,
              workers: int = 4, mode: str = "PU",
              progress_fn=None) -> list[dict]:
    """Evaluate all segments in parallel. Returns segments with results filled in.
    progress_fn(done, total, severity, score, category) is called after each segment.
    """
    total = len(segments)
    lock = threading.Lock()
    done = [0]

    _DASH_ONLY = {"-", "–", "—"}

    def evaluate_one(seg: dict) -> None:
        # Dash-only targets are intentional placeholders (non-English source workaround) — skip
        if seg.get("mt_target", "").strip() in _DASH_ONLY:
            seg["score"]          = 99
            seg["severity"]       = "OK"
            seg["error_category"] = "no-error"
            seg["reasoning"]      = ""
            with lock:
                done[0] += 1
                if progress_fn:
                    progress_fn(done[0], total, "OK", 99, "no-error")
            return

        try:
            if mode == "PE":
                result = evaluate_segment_pe(client, seg["source"], seg["mt_target"], termbase)
                seg["score"]          = None
                seg["severity"]       = "FAIL" if result.get("send_to_pe") == "Yes" else "OK"
                seg["error_category"] = result.get("improvement_type", "")
                seg["reasoning"]      = result.get("reasoning", "")
            else:
                result = _evaluate_with_retry(client, seg["source"], seg["mt_target"], termbase)
                seg["score"]          = result["score"]
                seg["severity"]       = result["severity"]
                seg["error_category"] = result.get("error_category", "")
                seg["reasoning"]      = result.get("reasoning", "")
        except Exception as exc:
            seg["score"]          = -1
            seg["severity"]       = "FAIL"
            seg["error_category"] = "api-error"
            seg["reasoning"]      = f"Segment skipped after retries: {exc}"

        score_str = str(seg["score"]) if seg["score"] is not None else "---"
        with lock:
            done[0] += 1
            print(f"  [{done[0]:>3}/{total}] {seg['severity']:4} ({score_str:>3}) | {seg['error_category']}")
            if progress_fn:
                progress_fn(done[0], total, seg["severity"], seg["score"], seg.get("error_category", ""))

    with ThreadPoolExecutor(max_workers=workers) as executor:
        futures = [executor.submit(evaluate_one, seg) for seg in segments]
        for future in as_completed(futures):
            future.result()

    return segments


def aggregate_by_validator(segments: list[dict]) -> dict:
    """
    Compute per-validator summary:
    - weighted average score (by segment, equal weight for now)
    - count by severity
    - count by error_category
    """
    from collections import defaultdict, Counter

    validators = defaultdict(lambda: {
        "scores": [],
        "severity_counts": Counter(),
        "category_counts": Counter(),
        "flagged_segments": [],
    })

    for seg in segments:
        v = seg["validator_name"]
        score = seg["score"]
        if score is not None and score >= 0:
            validators[v]["scores"].append(score)
        validators[v]["severity_counts"][seg["severity"]] += 1
        validators[v]["category_counts"][seg["error_category"]] += 1
        if seg["severity"] in ("WARN", "FAIL"):
            validators[v]["flagged_segments"].append(seg)

    summary = {}
    for v, data in validators.items():
        scores = data["scores"]
        avg_score = round(sum(scores) / len(scores), 2) if scores else None
        summary[v] = {
            "validator_name":  v,
            "total_segments":  len(scores),
            "avg_score":       avg_score,
            "severity_counts": dict(data["severity_counts"]),
            "category_counts": dict(data["category_counts"]),
            "flagged_count":   len(data["flagged_segments"]),
            "flagged_segments": data["flagged_segments"],
        }

    return summary


# ─────────────────────────────────────────────
# CHECKPOINT 4: EXCEL OUTPUT
# ─────────────────────────────────────────────

# Plain-language names for error categories (used in the LL-facing performance sheet)
FRIENDLY_CATEGORY = {
    "accuracy:mistranslation":               "Mistranslations",
    "accuracy:omission":                     "Omitted content",
    "accuracy:addition":                     "Added content",
    "accuracy:untranslated":                 "Untranslated English",
    "fluency:grammar":                       "Grammar",
    "fluency:spelling":                      "Spelling",
    "fluency:typography":                    "Typography/punctuation",
    "fluency:unintelligible":                "Unintelligible output",
    "style:unidiomatic":                     "Unnatural phrasing",
    "style:company-style":                   "Term/style deviations",
    "locale-convention:number-format":       "Number formatting",
    "locale-convention:measurement-format":  "Unit/measurement formatting",
}


def rate_resource(questionable_pct: float) -> tuple[str, str]:
    """Map a questionable-decision rate (%) to an LL-friendly rating + cell colour."""
    if questionable_pct < 5:
        return "Strong", "C6EFCE"                # green
    if questionable_pct < 15:
        return "Good", "DAF2D0"                  # pale green
    if questionable_pct < 30:
        return "Review recommended", "FFEB9C"    # yellow
    return "Attention needed", "FFC7CE"          # red


# Interpretation of the validator's decision, given the check mode and the
# evaluator's severity. Green = validator made the right call, red = wrong call.
def interpret_verdict(mode: str, severity: str, score=None) -> tuple[str, str]:
    """Return (verdict_text, hex_color) for a segment, given the check mode."""
    if mode == "PEQA":  # final QA of the linguist's post-edited string; gate at 95
        if score is not None and score < 95:
            return ("Below 95 — needs rework", "FFC7CE")       # red
        if severity == "WARN":
            return ("Borderline — quick check recommended", "FFEB9C")  # yellow (95–97)
        return ("Passes QA (>= 98)", "C6EFCE")                 # green
    if mode == "PU":  # validator chose to PUBLISH the string as-is
        return {
            "OK":   ("Correct publish — clean",            "C6EFCE"),  # green
            "WARN": ("Questionable — minor issues",        "FFEB9C"),  # yellow
            "FAIL": ("Wrong publish — has errors",         "FFC7CE"),  # red
        }.get(severity, ("—", "FFFFFF"))
    else:  # PE — validator chose to send the string to post-editing
        return {
            "OK":   ("Over-flagged — clean string sent to PE", "FFC7CE"),  # red
            "WARN": ("Borderline — minor issue",               "FFEB9C"),  # yellow
            "FAIL": ("Correct — real error, PE justified",     "C6EFCE"),  # green
        }.get(severity, ("—", "FFFFFF"))


def write_report(segments: list[dict], summary: dict, output_path: str | None,
                 mode: str = "PU", return_bytes: bool = False):
    """
    Write results to a separate Excel workbook with two sheets:
    - Segments: per-segment results
    - Resource_Performance: per-validator performance insight (mode-aware, LL-facing)
    """
    wb = openpyxl.Workbook()

    # ── Sheet 1: Segments ──
    ws_seg = wb.active
    ws_seg.title = "Segments"

    _LL_FEEDBACK_HEADERS = ["Agree with AI?", "If Not, Why?", "Feedback to Linguist"]

    if mode == "PEQA":
        seg_headers = [
            "SourceFile", "SegmentID", "Linguist", "GalileoID",
            "Source", "Post-edited target",
            "QA score", "Severity", "ErrorCategory", "Reasoning", "QA verdict",
        ] + _LL_FEEDBACK_HEADERS
    else:
        seg_headers = [
            "SourceFile", "SegmentID", "ValidatorName", "ValidatorID",
            "Source", "MT_Target",
            "Score", "Severity", "ErrorCategory", "Reasoning", "ValidatorVerdict",
        ] + _LL_FEEDBACK_HEADERS
    ws_seg.append(seg_headers)

    # Header styling
    from openpyxl.styles import Font, PatternFill, Alignment
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="2E4057")
    # Distinct teal fill for the LL input columns (L, M, N)
    ll_header_fill = PatternFill("solid", fgColor="1D7874")
    for col_idx, cell in enumerate(ws_seg[1], 1):
        cell.font = header_font
        cell.fill = ll_header_fill if col_idx >= 12 else header_fill
        cell.alignment = Alignment(horizontal="center")

    severity_colors = {
        "OK":   "C6EFCE",  # green
        "WARN": "FFEB9C",  # yellow
        "FAIL": "FFC7CE",  # red
    }

    for seg in segments:
        verdict_text, verdict_color = interpret_verdict(mode, seg["severity"], seg["score"])
        row = [
            seg["source_file"],
            seg["segment_id"],
            seg["validator_name"],
            seg["validator_id"],
            seg["source"],
            seg["mt_target"],
            seg["score"],
            seg["severity"],
            seg["error_category"],
            seg["reasoning"],
            verdict_text,
            "", "", "",  # L, M, N — LL fills in: Agree with AI? / If Not, Why? / Feedback to Linguist
        ]
        ws_seg.append(row)
        cur = ws_seg.max_row
        # Wrap text + vertical alignment for Source (E=5), Target (F=6), Reasoning (J=10)
        for col in (5, 6, 10):
            ws_seg.cell(row=cur, column=col).alignment = Alignment(wrap_text=True, vertical="top")
        ws_seg.row_dimensions[cur].height = 60
        # Colour the Severity cell
        sev_cell = ws_seg.cell(row=cur, column=8)
        sev_cell.fill = PatternFill("solid", fgColor=severity_colors.get(seg["severity"], "FFFFFF"))
        # Colour the ValidatorVerdict cell (green = right call, red = wrong call)
        verdict_cell = ws_seg.cell(row=cur, column=11)
        verdict_cell.fill = PatternFill("solid", fgColor=verdict_color)
        # Pre-apply filter: hide clean rows so only flagged/borderline are visible by default
        if mode == "PEQA" and verdict_text == "Passes QA (>= 98)":
            ws_seg.row_dimensions[cur].hidden = True

    # Column widths
    col_widths = [30, 45, 22, 12, 60, 60, 8, 10, 35, 80, 36, 16, 45, 45]
    for i, w in enumerate(col_widths, 1):
        ws_seg.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    # AutoFilter on header row + pre-set filter to show only flagged/borderline rows
    last_col = openpyxl.utils.get_column_letter(len(seg_headers))
    ws_seg.auto_filter.ref = f"A1:{last_col}1"
    if mode == "PEQA":
        try:
            from openpyxl.worksheet.filters import FilterColumn, Filters  # type: ignore[import]
            fc = FilterColumn(colId=10)  # K is 11th column; colId is 0-indexed from A
            fc.filters = Filters(filter=["Below 95 — needs rework", "Borderline — quick check recommended"])
            ws_seg.auto_filter.filterColumn.append(fc)
        except ImportError:
            pass  # hidden rows already handle filtering; dropdown arrows still appear

    # ── Sheet 2: Resource Performance (LL-facing insight, mode-aware) ──
    ws_sum = wb.create_sheet("Resource_Performance")

    below95_by_resource = {}
    if mode == "PEQA":
        explainer = ("PEQA — these are the linguists' POST-EDITED strings, scored for final quality. "
                     "'Below 95' = strings scoring under 95: flagged for the linguist to revisit and correct. "
                     "Lower below-95 rate = stronger post-editing quality.")
        questionable_sev = None
        for seg in segments:
            if seg.get("score") is not None and seg["score"] < 95:
                r = seg["validator_name"]
                below95_by_resource[r] = below95_by_resource.get(r, 0) + 1
    elif mode == "PU":
        explainer = ("PU check — these strings were PUBLISHED by the validator. "
                     "'Questionable' = strings the AI found errors in (FAIL): published despite errors. "
                     "WARN = borderline, not counted as questionable.")
        questionable_sev = "FAIL"
    else:
        explainer = ("PE check — these strings were SENT TO POST-EDITING by the validator. "
                     "'Questionable' = strings the AI found clean (OK): possibly unnecessary PE sends. "
                     "WARN = borderline, not counted as questionable.")
        questionable_sev = "OK"

    ws_sum.merge_cells("A1:I1")
    ws_sum["A1"] = explainer
    ws_sum["A1"].alignment = Alignment(wrap_text=True, vertical="top")
    ws_sum["A1"].font = Font(bold=True)
    ws_sum.merge_cells("A2:I2")
    ws_sum["A2"] = ("Note: based on automated AI evaluation — ratings are advisory and "
                    "intended as a starting point for Language Lead review, not a verdict.")
    ws_sum["A2"].font = Font(italic=True, size=9)

    sum_headers = [
        "Linguist" if mode == "PEQA" else "Validator", "Segments checked", "Avg score",
        "Clean (OK)", "Borderline (WARN)", "Errors found (FAIL)",
        "Below 95 rate" if mode == "PEQA" else "Questionable rate",
        "Overall rating", "Top issues found",
    ]
    ws_sum.append(sum_headers)
    for cell in ws_sum[3]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")

    for data in summary.values():
        sc = data["severity_counts"]
        n = data["total_segments"]
        if mode == "PEQA":
            questionable = below95_by_resource.get(data["validator_name"], 0)
        else:
            questionable = sc.get(questionable_sev, 0)
        q_rate = questionable / n * 100 if n else 0
        rating, rating_color = rate_resource(q_rate)

        top_cats = sorted(
            [(k, c) for k, c in data["category_counts"].items() if k not in ("no-error", None, "")],
            key=lambda x: -x[1]
        )[:3]
        top_cats_str = "; ".join(
            f"{FRIENDLY_CATEGORY.get(cat, cat)} ({c})" for cat, c in top_cats
        ) if top_cats else "—"

        ws_sum.append([
            data["validator_name"],
            n,
            data["avg_score"],
            sc.get("OK", 0),
            sc.get("WARN", 0),
            sc.get("FAIL", 0),
            f"{q_rate:.1f}%",
            rating,
            top_cats_str,
        ])
        ws_sum.cell(ws_sum.max_row, 8).fill = PatternFill("solid", fgColor=rating_color)
        ws_sum.cell(ws_sum.max_row, 8).font = Font(bold=True)

    sum_col_widths = [25, 16, 10, 11, 16, 16, 16, 20, 55]
    for i, w in enumerate(sum_col_widths, 1):
        ws_sum.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
    ws_sum.row_dimensions[1].height = 42
    ws_sum.row_dimensions[3].height = 28

    if return_bytes:
        import io as _io
        buf = _io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf
    wb.save(output_path)
    print(f"\n  Report saved: {output_path}")


# ─────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────


def ask_bearer_token() -> str:
    """Show a masked input dialog to collect the AWS Bedrock bearer token."""
    import tkinter as tk
    result = {"token": ""}

    win = tk.Tk()
    win.title("AWS Bedrock Token")
    win.resizable(False, False)
    win.attributes("-topmost", True)

    tk.Label(win, text="Enter your AWS Bedrock bearer token:").pack(padx=20, pady=(16, 4))
    entry = tk.Entry(win, show="*", width=56)
    entry.pack(padx=20, pady=4)
    entry.focus_set()

    def submit():
        result["token"] = re.sub(r"\s", "", entry.get())
        win.destroy()

    tk.Button(win, text="Continue", command=submit, width=12).pack(pady=(8, 16))
    win.bind("<Return>", lambda _: submit())
    win.protocol("WM_DELETE_WINDOW", win.destroy)
    win.mainloop()

    return result["token"]


def pick_files_dialog() -> list[str]:
    """Open a file picker dialog and return selected .xlsx paths."""
    import tkinter as tk
    from tkinter import filedialog
    root = tk.Tk()
    root.withdraw()
    root.attributes("-topmost", True)
    paths = filedialog.askopenfilenames(
        title="Select Galileo VA export file(s)",
        filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
    )
    root.destroy()
    return list(paths)


def ask_check_mode() -> str:
    """Ask which check to run. Returns 'PU', 'PE', or 'PEQA'."""
    import tkinter as tk
    result = {"mode": ""}

    win = tk.Tk()
    win.title("Check Mode")
    win.resizable(False, False)
    win.attributes("-topmost", True)

    tk.Label(win, text="Which check do you want to run?",
             font=("Segoe UI", 10, "bold")).pack(padx=24, pady=(18, 8))

    mode_var = tk.StringVar(value="PEQA")
    tk.Radiobutton(
        win, variable=mode_var, value="PEQA", justify="left", anchor="w",
        text="PEQA — Post-edit quality assurance\n"
             "Score the linguists' POST-EDITED output; flag anything below 95.",
    ).pack(fill="x", padx=24, pady=4)
    tk.Radiobutton(
        win, variable=mode_var, value="PU", justify="left", anchor="w",
        text="PU — Published strings (Publish = Yes)\n"
             "Check whether the validator was right to publish these as-is.",
    ).pack(fill="x", padx=24, pady=4)
    tk.Radiobutton(
        win, variable=mode_var, value="PE", justify="left", anchor="w",
        text="PE — Post-edited strings (Publish = No)\n"
             "Check whether the validator was right to send these to post-editing.",
    ).pack(fill="x", padx=24, pady=4)

    def submit():
        result["mode"] = mode_var.get()
        win.destroy()

    tk.Button(win, text="Continue", command=submit, width=12).pack(pady=(10, 18))
    win.bind("<Return>", lambda _: submit())
    win.protocol("WM_DELETE_WINDOW", win.destroy)
    win.mainloop()

    return result["mode"]


def ask_scope(total: int) -> int:
    """Ask full check (all rows) or spot check (random N). Returns number of rows to evaluate."""
    import tkinter as tk
    result = {"n": total}

    win = tk.Tk()
    win.title("Check Scope")
    win.resizable(False, False)
    win.attributes("-topmost", True)

    tk.Label(win, text=f"This file has {total} matching strings.",
             font=("Segoe UI", 10, "bold")).pack(padx=24, pady=(18, 8))

    scope_var = tk.StringVar(value="full")
    tk.Radiobutton(
        win, variable=scope_var, value="full", anchor="w", justify="left",
        text=f"Full check — evaluate all {total} strings.",
    ).pack(fill="x", padx=24, pady=(4, 2))

    spot_frame = tk.Frame(win)
    tk.Radiobutton(
        spot_frame, variable=scope_var, value="spot", anchor="w",
        text="Spot check — evaluate a random sample of:",
    ).pack(side="left")
    spot_entry = tk.Entry(spot_frame, width=6)
    spot_entry.insert(0, str(min(50, total)))
    spot_entry.pack(side="left", padx=(4, 0))
    spot_frame.pack(fill="x", padx=24, pady=(2, 4))

    def submit():
        if scope_var.get() == "full":
            result["n"] = total
        else:
            try:
                n = int(spot_entry.get())
            except (ValueError, TypeError):
                n = total
            result["n"] = max(1, min(n, total))
        win.destroy()

    tk.Button(win, text="Continue", command=submit, width=12).pack(pady=(10, 18))
    win.bind("<Return>", lambda _: submit())
    win.protocol("WM_DELETE_WINDOW", win.destroy)
    win.mainloop()

    return result["n"]


def main():
    if len(sys.argv) >= 2:
        input_files = sys.argv[1:]
    else:
        print("No files specified — opening file picker...")
        input_files = pick_files_dialog()
        if not input_files:
            print("No files selected. Exiting.")
            sys.exit(0)

    missing = [f for f in input_files if not Path(f).exists()]
    if missing:
        print(f"Error: file(s) not found: {missing}")
        sys.exit(1)

    # Token first, then check mode (asked after API key + file, per spec)
    token = ask_bearer_token()
    if not token:
        print("No token provided. Exiting.")
        sys.exit(1)

    mode = ask_check_mode()
    if not mode:
        print("No check mode selected. Exiting.")
        sys.exit(0)
    publish_filter = "Yes" if mode == "PU" else "No"
    if mode == "PEQA":
        mode_label = "PEQA — post-edited strings (final QA: is the linguist's output >= 95?)"
    elif mode == "PU":
        mode_label = "PU — published strings (was the validator right to publish?)"
    else:
        mode_label = "PE — post-edited strings (was the validator right to send to PE?)"

    print(f"\n=== VA Evaluator — Check mode: {mode_label} ===")
    print("\n=== Checkpoint 1: Reading files ===")
    if mode == "PEQA":
        segments = read_postedited_batch(input_files)
    else:
        segments = read_batch(input_files, publish_filter)
    print(f"\nTotal segments to evaluate: {len(segments)}")

    if not segments:
        msg = ("No post-edited segments found." if mode == "PEQA"
               else f"No Publish={publish_filter} segments found.")
        print(msg + " Exiting.")
        sys.exit(0)

    # Scope: full check or random spot check (asked after PU/PE selection)
    n_to_check = ask_scope(len(segments))
    if n_to_check < len(segments):
        segments = random.sample(segments, n_to_check)
        print(f"  Spot check: randomly selected {n_to_check} strings")
    else:
        print(f"  Full check: {len(segments)} strings")

    print("\n=== Checkpoint 2–3: Running AI evaluation ===")
    os.environ["AWS_BEARER_TOKEN_BEDROCK"] = token
    client = boto3.client(service_name="bedrock-runtime", region_name="us-east-1")
    termbase = load_termbase()
    print(f"  Termbase loaded: {len(termbase)} entries")
    segments = run_batch(segments, client, termbase, mode=mode)

    print("\n=== Checkpoint 3: Aggregating ===")
    summary = aggregate_by_validator(segments)
    for v, data in summary.items():
        print(f"  {v}: avg={data['avg_score']} | flagged={data['flagged_count']}/{data['total_segments']}")

    print("\n=== Checkpoint 4: Writing report ===")
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_path = f"va_report_{mode}_{timestamp}.xlsx"
    write_report(segments, summary, output_path, mode)

    print("\n=== Done ===")


if __name__ == "__main__":
    main()
