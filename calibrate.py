"""
calibrate.py
Measure evaluator agreement against the LL-filled goldset.

Usage:
    python calibrate.py            # full 77 answered rows
    python calibrate.py --limit 5  # first N rows (smoke test)
"""

import sys
import json
import argparse
import time
from datetime import datetime
from pathlib import Path
import os

import boto3
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment

from va_evaluator import evaluate_segment, BEDROCK_MODEL_ID, ask_bearer_token, load_termbase

GOLDSET_PATH = Path(__file__).parent / "repo" / "goldset_parsed.json"

TRUE_PASS  = "TRUE_PASS"   # LL Yes, eval OK
TRUE_FLAG  = "TRUE_FLAG"   # LL No,  eval WARN/FAIL
FALSE_PASS = "FALSE_PASS"  # LL No,  eval OK          ← dangerous: minimize
FALSE_FLAG = "FALSE_FLAG"  # LL Yes, eval WARN/FAIL   ← noisy but acceptable


def classify(ll_verdict: str, eval_severity: str) -> str:
    predicted_yes = eval_severity in ("OK", "WARN")  # WARN = advisory, still publishable
    ll_yes = ll_verdict == "Yes"
    if ll_yes and predicted_yes:
        return TRUE_PASS
    if not ll_yes and not predicted_yes:
        return TRUE_FLAG
    if not ll_yes and predicted_yes:
        return FALSE_PASS
    return FALSE_FLAG


def main():
    parser = argparse.ArgumentParser(description="Calibrate va_evaluator against LL goldset")
    parser.add_argument("--limit", type=int, default=None,
                        help="Evaluate only first N answered rows")
    args = parser.parse_args()

    rows = json.loads(GOLDSET_PATH.read_text(encoding="utf-8"))
    answered = [r for r in rows if r.get("verdict") is not None]
    if args.limit:
        answered = answered[:args.limit]
    total = len(answered)
    ll_no_total  = sum(1 for r in answered if r["verdict"] == "No")
    ll_yes_total = sum(1 for r in answered if r["verdict"] == "Yes")

    print(f"\n=== Calibration run: {total} rows (LL-Yes: {ll_yes_total}, LL-No: {ll_no_total}) ===")

    token = ask_bearer_token()
    if not token:
        print("No token provided. Exiting.")
        sys.exit(1)
    os.environ["AWS_BEARER_TOKEN_BEDROCK"] = token
    client = boto3.client(service_name="bedrock-runtime", region_name="us-east-1")
    termbase = load_termbase()
    print(f"  Termbase loaded: {len(termbase)} entries")

    run_start    = time.time()
    run_start_ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    results = []
    for i, row in enumerate(answered):
        print(f"  [{i+1}/{total}] row {row['row']:>3} | {row['source'][:55]}")
        t0 = time.time()
        eval_result = evaluate_segment(client, row["source"], row["mt"], termbase)
        proc_time_s = round(time.time() - t0, 1)
        label = classify(row["verdict"], eval_result["severity"])
        results.append({
            **row,
            "eval_severity":  eval_result["severity"],
            "eval_score":     eval_result["score"],
            "eval_category":  eval_result.get("error_category", ""),
            "eval_reasoning": eval_result.get("reasoning", ""),
            "proc_time_s":    proc_time_s,
            "input_tokens":   eval_result.get("input_tokens", 0),
            "output_tokens":  eval_result.get("output_tokens", 0),
            "latency_ms":     eval_result.get("latency_ms", 0),
            "tb_matches":     eval_result.get("tb_matches", 0),
            "classification": label,
        })
        flag = "  ← FALSE PASS" if label == FALSE_PASS else ""
        print(f"           → {eval_result['severity']:10} ({eval_result['score']:3}) | {label} | {proc_time_s}s{flag}")

    total_duration = round(time.time() - run_start, 1)

    # ── Metrics ──────────────────────────────────────────────────────────────
    counts = {TRUE_PASS: 0, TRUE_FLAG: 0, FALSE_PASS: 0, FALSE_FLAG: 0}
    for r in results:
        counts[r["classification"]] += 1

    agreement      = (counts[TRUE_PASS] + counts[TRUE_FLAG]) / total * 100
    recall_no      = counts[TRUE_FLAG] / ll_no_total * 100 if ll_no_total else 0
    n_flagged      = counts[TRUE_FLAG] + counts[FALSE_FLAG]
    precision_flag = counts[TRUE_FLAG] / n_flagged * 100 if n_flagged else 0

    print("\n" + "=" * 56)
    print(f"  Rows evaluated:      {total}")
    print(f"  Overall agreement:   {agreement:.1f}%")
    print(f"  TRUE_PASS:           {counts[TRUE_PASS]:3}  ({counts[TRUE_PASS]/total*100:.1f}%)")
    print(f"  TRUE_FLAG:           {counts[TRUE_FLAG]:3}  ({counts[TRUE_FLAG]/total*100:.1f}%)")
    print(f"  FALSE_PASS:          {counts[FALSE_PASS]:3}  ({counts[FALSE_PASS]/total*100:.1f}%)  ← minimize")
    print(f"  FALSE_FLAG:          {counts[FALSE_FLAG]:3}  ({counts[FALSE_FLAG]/total*100:.1f}%)")
    print(f"  Recall on No-rows:   {recall_no:.1f}%  (target ≥85%)")
    print(f"  Precision on flags:  {precision_flag:.1f}%")
    print("=" * 56)

    target_met = recall_no >= 85.0 and counts[FALSE_PASS] <= 6
    if target_met:
        print("\n  TARGET MET: recall >=85% and FALSE_PASS <=6")
    else:
        print("\n  TARGET MISSED — misses report below.")
        print("  Do NOT tune prompt automatically. Return results to Ahmet.")

    # ── FALSE_PASS detail (working list for next tuning pass) ────────────────
    false_passes = [r for r in results if r["classification"] == FALSE_PASS]
    if false_passes:
        print(f"\n── FALSE_PASS details ({len(false_passes)} rows) ──────────────────")
        for r in false_passes:
            print(f"\n  Row {r['row']}")
            print(f"  Source:    {r['source']}")
            print(f"  MT:        {r['mt']}")
            print(f"  LL reason: {r.get('reason', '—')}")
            print(f"  MT red:    {r.get('mt_red', [])}")
            print(f"  Eval:      {r['eval_severity']} ({r['eval_score']}) — {r['eval_reasoning']}")

    # ── Excel report ─────────────────────────────────────────────────────────
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_path = Path(__file__).parent / f"calibration_report_remap_{timestamp}.xlsx"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Calibration"

    headers = [
        "Row", "Source", "MT", "LL_verdict", "LL_reason",
        "MT_red_spans", "Eval_severity", "Eval_score",
        "Eval_category", "Eval_reasoning", "Classification",
        "Proc_time_s", "Input_tokens", "Output_tokens", "Latency_ms", "TB_matches",
    ]
    ws.append(headers)
    hdr_fill = PatternFill("solid", fgColor="2E4057")
    hdr_font = Font(bold=True, color="FFFFFF")
    for cell in ws[1]:
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = Alignment(horizontal="center")

    fill_map = {
        FALSE_PASS: PatternFill("solid", fgColor="FFC7CE"),  # red
        FALSE_FLAG: PatternFill("solid", fgColor="FFEB9C"),  # yellow
        TRUE_PASS:  PatternFill("solid", fgColor="C6EFCE"),  # green
        TRUE_FLAG:  PatternFill("solid", fgColor="C6EFCE"),  # green
    }

    for r in results:
        ws.append([
            r["row"],
            r["source"],
            r["mt"],
            r["verdict"],
            r.get("reason", ""),
            "; ".join(r.get("mt_red", [])),
            r["eval_severity"],
            r["eval_score"],
            r["eval_category"],
            r["eval_reasoning"],
            r["classification"],
            r.get("proc_time_s"),
            r.get("input_tokens"),
            r.get("output_tokens"),
            r.get("latency_ms"),
            r.get("tb_matches"),
        ])
        row_fill = fill_map.get(r["classification"])
        if row_fill:
            for cell in ws[ws.max_row]:
                cell.fill = row_fill

    col_widths = [6, 60, 60, 12, 60, 40, 14, 10, 30, 80, 14, 12, 13, 14, 12, 10]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    # ── Run_Log sheet ─────────────────────────────────────────────────────────
    ws_log = wb.create_sheet("Run_Log")
    ws_log.column_dimensions["A"].width = 26
    ws_log.column_dimensions["B"].width = 32

    proc_times          = [r["proc_time_s"] for r in results if r.get("proc_time_s") is not None]
    total_input_tokens  = sum(r.get("input_tokens", 0) for r in results)
    total_output_tokens = sum(r.get("output_tokens", 0) for r in results)

    log_entries = [
        ("Run timestamp",        run_start_ts),
        ("Model ID",             BEDROCK_MODEL_ID),
        ("Termbase entries",     len(termbase)),
        ("Rows evaluated",       total),
        ("Total duration (s)",   total_duration),
        ("Avg time/segment (s)", round(sum(proc_times) / len(proc_times), 1) if proc_times else 0),
        ("Min time/segment (s)", min(proc_times) if proc_times else 0),
        ("Max time/segment (s)", max(proc_times) if proc_times else 0),
        ("Total input tokens",   total_input_tokens),
        ("Total output tokens",  total_output_tokens),
        (None, None),
        ("TRUE_PASS",            counts[TRUE_PASS]),
        ("TRUE_FLAG",            counts[TRUE_FLAG]),
        ("FALSE_PASS",           counts[FALSE_PASS]),
        ("FALSE_FLAG",           counts[FALSE_FLAG]),
        ("Overall agreement",    f"{agreement:.1f}%"),
        ("Recall on No-rows",    f"{recall_no:.1f}%  (target >=85%)"),
        ("Precision on flags",   f"{precision_flag:.1f}%"),
        ("Target met",           "YES" if target_met else "NO"),
    ]

    key_font  = Font(bold=True)
    pass_fill = PatternFill("solid", fgColor="C6EFCE")
    fail_fill = PatternFill("solid", fgColor="FFC7CE")

    for key, val in log_entries:
        ws_log.append([key, val] if key else [])
        if key == "Target met":
            ws_log.cell(row=ws_log.max_row, column=2).fill = pass_fill if target_met else fail_fill
        if key:
            ws_log.cell(row=ws_log.max_row, column=1).font = key_font

    wb.save(output_path)
    print(f"\n  Report saved: {output_path}")

    # ── Markdown export (mirror of the workbook) ──────────────────────────────
    md_path = output_path.with_suffix(".md")

    def md_cell(val) -> str:
        return str(val if val is not None else "").replace("|", "\\|").replace("\n", " ").strip()

    md_lines = [
        "## Calibration",
        "| " + " | ".join(headers) + " |",
        "| " + " | ".join("---" for _ in headers) + " |",
    ]
    for r in results:
        row_vals = [
            r["row"], r["source"], r["mt"], r["verdict"], r.get("reason", ""),
            "; ".join(r.get("mt_red", [])), r["eval_severity"], r["eval_score"],
            r["eval_category"], r["eval_reasoning"], r["classification"],
            r.get("proc_time_s"), r.get("input_tokens"), r.get("output_tokens"),
            r.get("latency_ms"), r.get("tb_matches"),
        ]
        md_lines.append("| " + " | ".join(md_cell(v) for v in row_vals) + " |")

    md_lines += ["", "## Run_Log", "| Metric | Value |", "| --- | --- |"]
    for key, val in log_entries:
        if key:
            md_lines.append(f"| {md_cell(key)} | {md_cell(val)} |")

    md_path.write_text("\n".join(md_lines) + "\n", encoding="utf-8")
    print(f"  Report saved: {md_path}")

    if not target_met:
        sys.exit(1)


if __name__ == "__main__":
    main()
