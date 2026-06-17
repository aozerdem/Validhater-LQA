"""
calibrate_pe.py
Measure the PE-mode evaluator (PE_SYSTEM_PROMPT) against the linguist-labelled
PE goldset (repo/pe_goldset_parsed.json, 50 rows).

Ground truth: Birgitte's "Would linguist have sent to PE?" verdicts
(POST EDITING FEEDBACK_PE_va_report_PE_20260612_102519).

Baseline to beat — the old publish-calibrated prompt on these same 50 rows:
    recall on 'needs PE'   13/26 = 50.0%
    specificity on 'keep'  23/24 = 95.8%

Provisional targets: recall >= 80%, specificity >= 85%.
If missed: do NOT tune the prompt automatically. Return results to Ahmet.

Usage:
    python calibrate_pe.py            # full 50 rows
    python calibrate_pe.py --limit 5  # smoke test
"""

import sys
import json
import argparse
import time
from datetime import datetime
from pathlib import Path
import os

import boto3
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment

from va_evaluator import evaluate_segment_pe, BEDROCK_MODEL_ID, ask_bearer_token, load_termbase

GOLDSET_PATH = Path(__file__).parent / "repo" / "pe_goldset_parsed.json"

# Baseline (old publish prompt, FAIL => send) on this exact 50-row set
BASELINE_RECALL = 50.0
BASELINE_SPEC   = 95.8

TRUE_SEND   = "TRUE_SEND"    # LL Yes, AI Yes — real improvement caught
TRUE_KEEP   = "TRUE_KEEP"    # LL No,  AI No  — correctly left alone
MISSED_SEND = "MISSED_SEND"  # LL Yes, AI No  — dangerous: minimize
OVER_SEND   = "OVER_SEND"    # LL No,  AI Yes — noisy: erodes specificity
PARSE_ERROR = "PARSE_ERROR"


def classify(ll_verdict: str, ai_verdict: str) -> str:
    if ai_verdict not in ("Yes", "No"):
        return PARSE_ERROR
    if ll_verdict == "Yes":
        return TRUE_SEND if ai_verdict == "Yes" else MISSED_SEND
    return TRUE_KEEP if ai_verdict == "No" else OVER_SEND


def baseline_pred(prior_sev: str) -> str:
    """What the old publish prompt predicted: FAIL => send to PE."""
    return "Yes" if prior_sev == "FAIL" else "No"


def main():
    parser = argparse.ArgumentParser(description="Calibrate PE-mode evaluator against linguist goldset")
    parser.add_argument("--limit", type=int, default=None, help="Evaluate only first N rows")
    args = parser.parse_args()

    rows = json.loads(GOLDSET_PATH.read_text(encoding="utf-8"))
    if args.limit:
        rows = rows[:args.limit]
    total = len(rows)
    ll_yes_total = sum(1 for r in rows if r["ll_send_to_pe"] == "Yes")
    ll_no_total  = sum(1 for r in rows if r["ll_send_to_pe"] == "No")

    print(f"\n=== PE calibration run: {total} rows "
          f"(LL send-to-PE: {ll_yes_total}, LL keep: {ll_no_total}) ===")
    print(f"    Baseline to beat: recall {BASELINE_RECALL}%, specificity {BASELINE_SPEC}%")

    token = ask_bearer_token()
    if not token:
        print("No token provided. Exiting.")
        sys.exit(1)
    os.environ["AWS_BEARER_TOKEN_BEDROCK"] = token
    client = boto3.client(service_name="bedrock-runtime", region_name="us-east-1")
    termbase = load_termbase()
    print(f"  Termbase loaded: {len(termbase)} entries")

    run_start    = time.time()
    run_start_ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    results = []
    for i, row in enumerate(rows):
        print(f"  [{i+1}/{total}] idx {row['idx']:>2} seg {row['seg']:>4} | {row['source'][:50]}")
        t0 = time.time()
        eval_result = evaluate_segment_pe(client, row["source"], row["mt"], termbase)
        proc_time_s = round(time.time() - t0, 1)
        label = classify(row["ll_send_to_pe"], eval_result["send_to_pe"])
        results.append({
            **row,
            "ai_send_to_pe":  eval_result["send_to_pe"],
            "ai_improvement": eval_result.get("improvement_type", ""),
            "ai_reasoning":   eval_result.get("reasoning", ""),
            "proc_time_s":    proc_time_s,
            "input_tokens":   eval_result.get("input_tokens", 0),
            "output_tokens":  eval_result.get("output_tokens", 0),
            "latency_ms":     eval_result.get("latency_ms", 0),
            "tb_matches":     eval_result.get("tb_matches", 0),
            "classification": label,
            "baseline_pred":  baseline_pred(row["prior_ai_sev"]),
        })
        flag = "  <- MISSED" if label == MISSED_SEND else ("  <- OVER-SEND" if label == OVER_SEND else "")
        print(f"           -> AI={eval_result['send_to_pe']:3} (LL={row['ll_send_to_pe']:3}) | {label} | {proc_time_s}s{flag}")

    total_duration = round(time.time() - run_start, 1)

    # ── Metrics ──────────────────────────────────────────────────────────────
    counts = {TRUE_SEND: 0, TRUE_KEEP: 0, MISSED_SEND: 0, OVER_SEND: 0, PARSE_ERROR: 0}
    for r in results:
        counts[r["classification"]] += 1

    recall      = counts[TRUE_SEND] / ll_yes_total * 100 if ll_yes_total else 0
    specificity = counts[TRUE_KEEP] / ll_no_total * 100 if ll_no_total else 0
    agreement   = (counts[TRUE_SEND] + counts[TRUE_KEEP]) / total * 100

    # Per-row comparison vs the old publish prompt
    fixed = sum(1 for r in results
                if r["ll_send_to_pe"] == "Yes" and r["baseline_pred"] == "No"
                and r["ai_send_to_pe"] == "Yes")
    lost = sum(1 for r in results
               if r["ll_send_to_pe"] == "Yes" and r["baseline_pred"] == "Yes"
               and r["ai_send_to_pe"] != "Yes")
    new_oversend = sum(1 for r in results
                       if r["ll_send_to_pe"] == "No" and r["baseline_pred"] == "No"
                       and r["ai_send_to_pe"] == "Yes")

    print("\n" + "=" * 60)
    print(f"  Rows evaluated:        {total}")
    print(f"  TRUE_SEND:             {counts[TRUE_SEND]:3}")
    print(f"  TRUE_KEEP:             {counts[TRUE_KEEP]:3}")
    print(f"  MISSED_SEND:           {counts[MISSED_SEND]:3}  <- minimize")
    print(f"  OVER_SEND:             {counts[OVER_SEND]:3}")
    print(f"  PARSE_ERROR:           {counts[PARSE_ERROR]:3}")
    print(f"  Recall on 'needs PE':  {recall:.1f}%   (baseline {BASELINE_RECALL}%, target >=80%)")
    print(f"  Specificity on 'keep': {specificity:.1f}%   (baseline {BASELINE_SPEC}%, target >=85%)")
    print(f"  Overall agreement:     {agreement:.1f}%   (note: 'keep' class enriched vs production)")
    print(f"  vs old prompt: fixed {fixed} misses | lost {lost} catches | {new_oversend} new over-sends")
    print("=" * 60)

    target_met = recall >= 80.0 and specificity >= 85.0
    if target_met:
        print("\n  TARGET MET: recall >=80% and specificity >=85%")
    else:
        print("\n  TARGET MISSED.")
        print("  Do NOT tune the prompt automatically. Return results to Ahmet.")

    # ── Miss details ─────────────────────────────────────────────────────────
    for label, title in ((MISSED_SEND, "MISSED_SEND (AI kept, linguist would PE)"),
                         (OVER_SEND, "OVER_SEND (AI sends, linguist would keep)")):
        bad = [r for r in results if r["classification"] == label]
        if bad:
            print(f"\n-- {title}: {len(bad)} --")
            for r in bad:
                safe_src = r["source"][:60].encode("ascii", "replace").decode()
                safe_rea = str(r["ai_reasoning"])[:120].encode("ascii", "replace").decode()
                print(f"  idx {r['idx']} seg {r['seg']} | old={r['prior_ai_sev']} {r['prior_ai_score']} | {safe_src}")
                print(f"      AI: {safe_rea}")

    # ── Excel report ─────────────────────────────────────────────────────────
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_path = Path(__file__).parent / f"pe_calibration_{timestamp}.xlsx"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "PE_Calibration"

    headers = [
        "Idx", "Seg", "Source", "MT", "LL_send_to_PE",
        "Old_prompt_sev", "Old_prompt_pred", "AI_send_to_PE",
        "Improvement_type", "AI_reasoning", "Classification",
        "Proc_time_s", "Input_tokens", "Output_tokens", "Latency_ms", "TB_matches",
    ]
    ws.append(headers)
    hdr_fill = PatternFill("solid", fgColor="2E4057")
    hdr_font = Font(bold=True, color="FFFFFF")
    for cell in ws[1]:
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = Alignment(horizontal="center")

    fill_map = {
        MISSED_SEND: PatternFill("solid", fgColor="FFC7CE"),  # red
        OVER_SEND:   PatternFill("solid", fgColor="FFEB9C"),  # yellow
        PARSE_ERROR: PatternFill("solid", fgColor="D9D2E9"),  # purple
        TRUE_SEND:   PatternFill("solid", fgColor="C6EFCE"),  # green
        TRUE_KEEP:   PatternFill("solid", fgColor="C6EFCE"),  # green
    }

    for r in results:
        ws.append([
            r["idx"], r["seg"], r["source"], r["mt"], r["ll_send_to_pe"],
            r["prior_ai_sev"], r["baseline_pred"], r["ai_send_to_pe"],
            r["ai_improvement"], r["ai_reasoning"], r["classification"],
            r.get("proc_time_s"), r.get("input_tokens"), r.get("output_tokens"),
            r.get("latency_ms"), r.get("tb_matches"),
        ])
        row_fill = fill_map.get(r["classification"])
        if row_fill:
            for cell in ws[ws.max_row]:
                cell.fill = row_fill

    col_widths = [6, 8, 55, 55, 14, 14, 16, 14, 28, 80, 14, 12, 13, 14, 12, 10]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    # ── Run_Log sheet ─────────────────────────────────────────────────────────
    ws_log = wb.create_sheet("Run_Log")
    ws_log.column_dimensions["A"].width = 28
    ws_log.column_dimensions["B"].width = 40

    proc_times          = [r["proc_time_s"] for r in results if r.get("proc_time_s") is not None]
    total_input_tokens  = sum(r.get("input_tokens", 0) for r in results)
    total_output_tokens = sum(r.get("output_tokens", 0) for r in results)

    log_entries = [
        ("Run timestamp",          run_start_ts),
        ("Model ID",               BEDROCK_MODEL_ID),
        ("Prompt",                 "PE_SYSTEM_PROMPT (post-editing triage)"),
        ("Termbase entries",       len(termbase)),
        ("Rows evaluated",         total),
        ("Total duration (s)",     total_duration),
        ("Avg time/segment (s)",   round(sum(proc_times) / len(proc_times), 1) if proc_times else 0),
        ("Total input tokens",     total_input_tokens),
        ("Total output tokens",    total_output_tokens),
        (None, None),
        ("TRUE_SEND",              counts[TRUE_SEND]),
        ("TRUE_KEEP",              counts[TRUE_KEEP]),
        ("MISSED_SEND",            counts[MISSED_SEND]),
        ("OVER_SEND",              counts[OVER_SEND]),
        ("PARSE_ERROR",            counts[PARSE_ERROR]),
        ("Recall on 'needs PE'",   f"{recall:.1f}%  (baseline {BASELINE_RECALL}%, target >=80%)"),
        ("Specificity on 'keep'",  f"{specificity:.1f}%  (baseline {BASELINE_SPEC}%, target >=85%)"),
        ("Overall agreement",      f"{agreement:.1f}%  (keep class enriched vs production)"),
        ("Fixed old misses",       f"{fixed} of 13"),
        ("Lost old catches",       lost),
        ("New over-sends",         new_oversend),
        ("Target met",             "YES" if target_met else "NO"),
    ]

    key_font  = Font(bold=True)
    pass_fill = PatternFill("solid", fgColor="C6EFCE")
    fail_fill = PatternFill("solid", fgColor="FFC7CE")

    for key, val in log_entries:
        ws_log.append([key, val] if key else [])
        if key == "Target met":
            ws_log.cell(row=ws_log.max_row, column=2).fill = pass_fill if target_met else fail_fill
        if key:
            ws_log.cell(row=ws_log.max_row, column=1).font = key_font

    wb.save(output_path)
    print(f"\n  Report saved: {output_path}")

    # ── Markdown export (mirror of the workbook) ──────────────────────────────
    md_path = output_path.with_suffix(".md")

    def md_cell(val) -> str:
        return str(val if val is not None else "").replace("|", "\\|").replace("\n", " ").strip()

    md_lines = [
        "## PE_Calibration",
        "| " + " | ".join(headers) + " |",
        "| " + " | ".join("---" for _ in headers) + " |",
    ]
    for r in results:
        row_vals = [
            r["idx"], r["seg"], r["source"], r["mt"], r["ll_send_to_pe"],
            r["prior_ai_sev"], r["baseline_pred"], r["ai_send_to_pe"],
            r["ai_improvement"], r["ai_reasoning"], r["classification"],
            r.get("proc_time_s"), r.get("input_tokens"), r.get("output_tokens"),
            r.get("latency_ms"), r.get("tb_matches"),
        ]
        md_lines.append("| " + " | ".join(md_cell(v) for v in row_vals) + " |")

    md_lines += ["", "## Run_Log", "| Metric | Value |", "| --- | --- |"]
    for key, val in log_entries:
        if key:
            md_lines.append(f"| {md_cell(key)} | {md_cell(val)} |")

    md_path.write_text("\n".join(md_lines) + "\n", encoding="utf-8")
    print(f"  Report saved: {md_path}")

    if not target_met:
        sys.exit(1)


if __name__ == "__main__":
    main()
