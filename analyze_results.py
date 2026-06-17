"""
analyze_results.py
Post-run analysis of va_evaluator reports — no human input required.

What it does:
  1. Stats summary       — severity/category distributions, per-validator profiles
  2. Pattern mining      — recurring issue types counted across reasonings
  3. Termbase audit      — which termbase entries triggered flags (blacklist candidates)
  4. Arbitration sheet   — disagreement rows pre-formatted for linguist verdicts
  5. AI meta-analysis    — optional (--ai): one Claude call over the aggregates

Usage:
    python analyze_results.py report1.xlsx report2.md ...
    python analyze_results.py --ai report1.xlsx        # adds AI insights sheet
    python analyze_results.py                          # opens file picker

Accepts both .xlsx reports (Segments sheet) and .md exports of the same.
Output: analysis_<timestamp>.xlsx
"""

import sys
import json
import re
import argparse
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path
import tkinter as tk
from tkinter import filedialog

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment


# ─────────────────────────────────────────────
# REPORT LOADING (.xlsx and .md)
# ─────────────────────────────────────────────

def load_report_xlsx(filepath: str) -> list[dict]:
    wb = openpyxl.load_workbook(filepath, data_only=True, read_only=True)
    sheet = "Segments" if "Segments" in wb.sheetnames else wb.sheetnames[0]
    ws = wb[sheet]
    rows_iter = ws.iter_rows(values_only=True)
    headers = [str(h) if h is not None else "" for h in next(rows_iter)]
    rows = []
    for raw in rows_iter:
        row = {h: ("" if v is None else v) for h, v in zip(headers, raw)}
        if any(str(v).strip() for v in row.values()):
            rows.append(row)
    wb.close()
    return rows


def load_report_md(filepath: str) -> list[dict]:
    """Parse the first markdown table in the file (the Segments export)."""
    lines = Path(filepath).read_text(encoding="utf-8").splitlines()
    headers, rows, in_table = None, [], False
    for line in lines:
        stripped = line.strip()
        if in_table and stripped.startswith("## "):
            break  # next sheet section — stop at the first table
        if not stripped.startswith("|"):
            continue
        cells = [c.strip().replace("\\_", "_") for c in stripped.split("|")[1:-1]]
        if all(set(c) <= {"-", " ", ":"} for c in cells):
            continue  # separator row
        if headers is None:
            headers = cells
            in_table = True
            continue
        cells = ["" if c == "NaN" else c for c in cells]
        rows.append({h: (cells[i] if i < len(cells) else "") for i, h in enumerate(headers)})
    return rows


def load_report(filepath: str) -> list[dict]:
    if filepath.lower().endswith(".md"):
        return load_report_md(filepath)
    return load_report_xlsx(filepath)


def detect_mode(rows: list[dict], filepath: str) -> str:
    """Return 'PU' or 'PE' from the ValidatorVerdict column, falling back to filename."""
    verdicts = " ".join(str(r.get("ValidatorVerdict", "")) for r in rows[:80])
    if "sent to PE" in verdicts or "PE justified" in verdicts:
        return "PE"
    if "publish" in verdicts.lower():
        return "PU"
    name = Path(filepath).name.upper()
    if name.startswith("PE") or "_PE_" in name:
        return "PE"
    return "PU"


def to_int(val) -> int | None:
    try:
        return int(float(val))
    except (ValueError, TypeError):
        return None


# ─────────────────────────────────────────────
# 2. PATTERN MINING
# ─────────────────────────────────────────────

PATTERNS = [
    ("Termbase-cited flag",            r"termbase"),
    ("'vennligst' unidiomatic",        r"vennligst"),
    ("Number+unit spacing",            r"missing space|number and unit"),
    ("Hyphen vs en-dash in ranges",    r"en-dash|en dash"),
    ("Title case / capitalisation",    r"title.?case|capitalis|sentence.?case"),
    ("Tag/markup placement",           r"\btags?\b|placeholder"),
    ("Calque / literal translation",   r"calque|literal"),
    ("Untranslated English",           r"untranslated|left in english"),
    ("Decimal separator",              r"decimal"),
    ("Known-lexicon term hit",         r"baseballhett|bordskål|vakuumpos|\bflott\b|\bdrakt\b|\blue\b"),
    ("False friend / semantic drift",  r"false.?friend|semantic"),
    ("Grammar agreement",              r"agreement|gender|plural|definite"),
]


def mine_patterns(rows: list[dict]) -> Counter:
    counts = Counter()
    for r in rows:
        reasoning = str(r.get("Reasoning", "")).lower()
        if not reasoning:
            continue
        for label, pattern in PATTERNS:
            if re.search(pattern, reasoning):
                counts[label] += 1
    return counts


# ─────────────────────────────────────────────
# 3. TERMBASE AUDIT
# ─────────────────────────────────────────────

DISMISS_KEYWORDS = re.compile(
    r"not appli|doesn'?t apply|wrong.?domain|not relevant|correctly not|clearly not", re.I)
PAIR_RE = re.compile(
    r"['\"‘’«]?([A-Za-zÆØÅæøå][\w\s'-]{0,28}?)"
    r"['\"‘’»]?\s*(?:→|->)\s*"
    r"['\"‘’«]?([A-Za-zÆØÅæøå][\w\s'-]{0,28}?)"
    r"['\"‘’»]?(?=[\s,.;)]|$)")


def audit_termbase(rows: list[dict]) -> dict:
    """Collect termbase-citing flags: {(en, nb): {"cited": n, "dismissed": n, "example": src}}."""
    audit = defaultdict(lambda: {"cited": 0, "dismissed": 0, "example": ""})
    for r in rows:
        reasoning = str(r.get("Reasoning", ""))
        if "termbase" not in reasoning.lower():
            continue
        dismissed = bool(DISMISS_KEYWORDS.search(reasoning))
        pairs = []
        for a, b in PAIR_RE.findall(reasoning):
            if any(ch.isdigit() for ch in a + b):
                continue
            # Strip narration prefixes the regex may capture ("Termbase specifies 'X")
            a = re.sub(r"^.*?(?:specifies|requires|suggests|entry|the termbase)\s*['\"‘’«]?",
                       "", a, flags=re.I).strip(" '\"‘’«»")
            b = b.strip(" '\"‘’«»")
            if a and b:
                pairs.append((a.lower(), b.lower()))
        if not pairs:
            pairs = [("(unparsed)", reasoning[:60])]
        for pair in pairs:
            entry = audit[pair]
            entry["dismissed" if dismissed else "cited"] += 1
            if not entry["example"]:
                entry["example"] = str(r.get("Source", ""))[:80]
    return dict(audit)


# ─────────────────────────────────────────────
# 5. AI META-ANALYSIS (optional, one call)
# ─────────────────────────────────────────────

META_PROMPT = """You are reviewing aggregate results from an automated MT QA evaluator \
(EN-GB to Norwegian Bokmal, Amazon product data). You receive summary statistics, recurring \
error patterns, a termbase audit (entries the evaluator cited when flagging), and sample \
disagreement rows where the evaluator and the human validator disagree.

Provide concise, actionable advice in these sections:
1. TERMBASE BLACKLIST - entries that look wrong-domain for product copy and should be excluded, with one-line rationale each.
2. PROMPT IMPROVEMENTS - up to 5 specific changes to the evaluator's system prompt.
3. LEXICON CANDIDATES - recurring term errors worth adding to the known-error lexicon.
4. PRE-CHECK CANDIDATES - deterministic regex checks that could replace model judgement.
5. EVALUATOR BEHAVIOUR - anything suspicious (over-strictness, leniency, category confusion).

Be specific. Cite the data you were given. Plain text, no markdown tables."""


def run_ai_meta_analysis(aggregate: dict) -> str:
    import os
    import boto3
    from va_evaluator import ask_bearer_token, BEDROCK_MODEL_ID

    token = ask_bearer_token()
    if not token:
        return "(skipped — no token provided)"
    os.environ["AWS_BEARER_TOKEN_BEDROCK"] = token
    client = boto3.client(service_name="bedrock-runtime", region_name="us-east-1")

    response = client.converse(
        modelId=BEDROCK_MODEL_ID,
        system=[{"text": META_PROMPT}],
        messages=[{"role": "user", "content": [{"text": json.dumps(aggregate, ensure_ascii=False)}]}],
        inferenceConfig={"maxTokens": 2000},
    )
    return response["output"]["message"]["content"][0]["text"].strip()


# ─────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────

def pick_files() -> list[str]:
    root = tk.Tk()
    root.withdraw()
    root.attributes("-topmost", True)
    paths = filedialog.askopenfilenames(
        title="Select va_report file(s) to analyze",
        filetypes=[("Reports", "*.xlsx *.md"), ("All files", "*.*")],
    )
    root.destroy()
    return list(paths)


def main():
    parser = argparse.ArgumentParser(description="Analyze va_evaluator reports")
    parser.add_argument("reports", nargs="*", help="Report files (.xlsx or .md)")
    parser.add_argument("--ai", action="store_true", help="Run the AI meta-analysis pass")
    args = parser.parse_args()

    files = args.reports or pick_files()
    if not files:
        print("No files selected. Exiting.")
        sys.exit(0)
    missing = [f for f in files if not Path(f).exists()]
    if missing:
        print(f"Error: file(s) not found: {missing}")
        sys.exit(1)

    # ── Load all reports ─────────────────────────────────────────────────────
    all_rows = []          # (mode, row) tuples
    file_summaries = []    # per-file stats for the Summary sheet
    for fp in files:
        rows = load_report(fp)
        mode = detect_mode(rows, fp)
        print(f"  Loaded: {Path(fp).name}  ({len(rows)} rows, mode={mode})")
        all_rows.extend((mode, r) for r in rows)

        sev = Counter(str(r.get("Severity", "")) for r in rows)
        scores = [s for s in (to_int(r.get("Score")) for r in rows) if s is not None and s >= 0]
        n = len(rows) or 1
        file_summaries.append({
            "file": Path(fp).name, "mode": mode, "rows": len(rows),
            "ok": sev.get("OK", 0), "warn": sev.get("WARN", 0), "fail": sev.get("FAIL", 0),
            "ok_pct": sev.get("OK", 0) / n * 100,
            "warn_pct": sev.get("WARN", 0) / n * 100,
            "fail_pct": sev.get("FAIL", 0) / n * 100,
            "avg_score": round(sum(scores) / len(scores), 1) if scores else None,
        })

    rows_flat = [r for _, r in all_rows]

    # ── 1. Stats ─────────────────────────────────────────────────────────────
    categories = Counter(str(r.get("ErrorCategory", "")) for r in rows_flat
                         if str(r.get("ErrorCategory", "")) not in ("", "no-error"))
    validators = defaultdict(lambda: Counter())
    for mode, r in all_rows:
        validators[str(r.get("ValidatorName", "?"))][f"{mode}:{r.get('Severity', '')}"] += 1

    # ── 2. Patterns ──────────────────────────────────────────────────────────
    patterns = mine_patterns(rows_flat)

    # ── 3. Termbase audit ────────────────────────────────────────────────────
    tb_audit = audit_termbase(rows_flat)

    # ── 4. Disagreements (arbitration candidates) ────────────────────────────
    disagreements = []
    for mode, r in all_rows:
        sev = str(r.get("Severity", ""))
        if mode == "PE" and sev == "OK":
            dtype = "PE-OK: AI says clean, validator sent to PE"
        elif mode == "PU" and sev == "FAIL":
            dtype = "PU-FAIL: AI says error, validator published"
        elif sev == "WARN":
            dtype = "Borderline (WARN)"
        else:
            continue
        disagreements.append((mode, dtype, r))
    hard = [d for d in disagreements if not d[1].startswith("Borderline")]
    print(f"\n  Disagreements: {len(hard)} hard, "
          f"{len(disagreements) - len(hard)} borderline")

    # ── Console summary ──────────────────────────────────────────────────────
    print("\n" + "=" * 60)
    for fs in file_summaries:
        print(f"  {fs['file']}  [{fs['mode']}]")
        print(f"    OK {fs['ok']} ({fs['ok_pct']:.0f}%) | WARN {fs['warn']} ({fs['warn_pct']:.0f}%)"
              f" | FAIL {fs['fail']} ({fs['fail_pct']:.0f}%) | avg score {fs['avg_score']}")
        if fs["mode"] == "PE":
            print(f"    -> {fs['ok_pct']:.0f}% of PE'd strings look clean to the AI (potential over-flagging by validator)")
        else:
            print(f"    -> {fs['fail_pct']:.0f}% of published strings look bad to the AI (potential wrong publishes)")
    print("\n  Top error categories:")
    for cat, n in categories.most_common(8):
        print(f"    {n:>4}  {cat}")
    print("\n  Recurring patterns:")
    for label, n in patterns.most_common():
        print(f"    {n:>4}  {label}")
    if tb_audit:
        print("\n  Termbase entries cited in flags (review for blacklist):")
        for (en, nb), d in sorted(tb_audit.items(), key=lambda x: -(x[1]['cited'] + x[1]['dismissed'])):
            note = "  <- model itself dismissed as wrong-domain" if d["dismissed"] else ""
            print(f"    {d['cited'] + d['dismissed']:>3}x  {en} -> {nb}{note}")
    print("=" * 60)

    # ── 5. AI meta-analysis (optional) ───────────────────────────────────────
    ai_text = ""
    if args.ai:
        print("\n  Running AI meta-analysis (single call)...")
        sample = [{
            "mode": m, "type": t,
            "source": str(r.get("Source", ""))[:150],
            "mt": str(r.get("MT_Target", ""))[:150],
            "score": r.get("Score"), "category": r.get("ErrorCategory"),
            "reasoning": str(r.get("Reasoning", ""))[:200],
        } for m, t, r in hard[:15]]
        aggregate = {
            "file_summaries": file_summaries,
            "top_error_categories": dict(categories.most_common(10)),
            "recurring_patterns": dict(patterns),
            "termbase_audit": {f"{en} -> {nb}": d for (en, nb), d in tb_audit.items()},
            "sample_disagreements": sample,
        }
        ai_text = run_ai_meta_analysis(aggregate)
        print("\n" + ai_text)

    # ── Excel output ─────────────────────────────────────────────────────────
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    out_path = Path(__file__).parent / f"analysis_{timestamp}.xlsx"
    wb = openpyxl.Workbook()
    hdr_fill = PatternFill("solid", fgColor="2E4057")
    hdr_font = Font(bold=True, color="FFFFFF")

    def style_header(ws):
        for cell in ws[1]:
            cell.fill = hdr_fill
            cell.font = hdr_font
            cell.alignment = Alignment(horizontal="center")

    # Summary
    ws = wb.active
    ws.title = "Summary"
    ws.append(["File", "Mode", "Rows", "OK", "WARN", "FAIL",
               "OK %", "WARN %", "FAIL %", "Avg score"])
    for fs in file_summaries:
        ws.append([fs["file"], fs["mode"], fs["rows"], fs["ok"], fs["warn"], fs["fail"],
                   round(fs["ok_pct"], 1), round(fs["warn_pct"], 1),
                   round(fs["fail_pct"], 1), fs["avg_score"]])
    ws.append([])
    ws.append(["Error category", "Count"])
    for cat, n in categories.most_common():
        ws.append([cat, n])
    style_header(ws)
    for i, w in enumerate([42, 8, 8, 8, 8, 8, 8, 9, 9, 10], 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    # Patterns
    ws = wb.create_sheet("Patterns")
    ws.append(["Pattern", "Occurrences"])
    for label, n in patterns.most_common():
        ws.append([label, n])
    style_header(ws)
    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 14

    # Termbase audit
    ws = wb.create_sheet("Termbase_Audit")
    ws.append(["EN term", "NB term", "Cited in flags", "Dismissed as wrong-domain",
               "Example source", "Suggested action"])
    warn_fill = PatternFill("solid", fgColor="FFC7CE")
    for (en, nb), d in sorted(tb_audit.items(), key=lambda x: -(x[1]["cited"] + x[1]["dismissed"])):
        action = "BLACKLIST candidate" if d["dismissed"] else "Review with linguist"
        ws.append([en, nb, d["cited"], d["dismissed"], d["example"], action])
        if d["dismissed"]:
            for cell in ws[ws.max_row]:
                cell.fill = warn_fill
    style_header(ws)
    for i, w in enumerate([24, 24, 14, 22, 60, 22], 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    # Arbitration
    ws = wb.create_sheet("Arbitration")
    ws.append(["Mode", "Disagreement type", "SegmentID", "ValidatorName", "Source", "MT_Target",
               "Score", "Severity", "ErrorCategory", "AI reasoning",
               "Linguist verdict (AI right / Validator right / Both partly)", "Linguist notes"])
    hard_fill = PatternFill("solid", fgColor="FFC7CE")
    soft_fill = PatternFill("solid", fgColor="FFEB9C")
    for mode, dtype, r in sorted(disagreements, key=lambda d: d[1].startswith("Borderline")):
        ws.append([mode, dtype, r.get("SegmentID", ""), r.get("ValidatorName", ""),
                   r.get("Source", ""), r.get("MT_Target", ""),
                   r.get("Score", ""), r.get("Severity", ""), r.get("ErrorCategory", ""),
                   r.get("Reasoning", ""), "", ""])
        fill = soft_fill if dtype.startswith("Borderline") else hard_fill
        ws.cell(row=ws.max_row, column=2).fill = fill
    style_header(ws)
    for i, w in enumerate([6, 40, 12, 18, 55, 55, 8, 10, 28, 70, 30, 40], 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    # AI insights
    if ai_text:
        ws = wb.create_sheet("AI_Insights")
        ws.column_dimensions["A"].width = 120
        ws.append(["AI meta-analysis"])
        style_header(ws)
        for line in ai_text.splitlines():
            ws.append([line])

    wb.save(out_path)
    print(f"\n  Analysis saved: {out_path}")


if __name__ == "__main__":
    main()
